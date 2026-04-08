"""
送迎ルート自動作成システム - Streamlit Webアプリ v7
=====================================================
放課後等デイサービス / 就労継続支援A型・B型 対応

v7 改修内容:
  1. VRP Time Window のシンプル化（4項目のみ）
     早迎え禁止時刻 / 迎えリミット / 送り出発可能時刻 / 送り到着リミット
  2. 利用者マスタに「基本利用曜日」（月〜日）列を追加
     カレンダーは曜日×マスタ時間を参照するExcel数式で自動生成
  3. カレンダー連動の自動チェック + UI上での直前微調整
"""

from __future__ import annotations

import io
import math
import re
import datetime
import calendar as cal_mod
from dataclasses import dataclass, field
from typing import Optional
from enum import Enum

import pandas as pd
import streamlit as st

# ---- オプションライブラリ ----
try:
    import folium
    from streamlit_folium import st_folium
    FOLIUM_AVAILABLE = True
except ImportError:
    FOLIUM_AVAILABLE = False

try:
    from ortools.constraint_solver import routing_enums_pb2, pywrapcp
    ORTOOLS_AVAILABLE = True
except ImportError:
    ORTOOLS_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ==============================================================
# ページ設定
# ==============================================================
st.set_page_config(
    page_title="送迎ルート最適化 v7",
    page_icon="🚌",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ==============================================================
# テーマ定数
# ==============================================================
THEME = {
    "primary":    "#1B3A5C",
    "primary_lt": "#2D5F8A",
    "accent":     "#2E7D52",
    "accent_lt":  "#E8F5EE",
    "warn":       "#C0392B",
    "warn_lt":    "#FDECEA",
    "surface":    "#FFFFFF",
    "surface2":   "#F6F8FA",
    "border":     "#DDE1E7",
    "text":       "#1C2330",
    "text2":      "#5A6478",
}

# ==============================================================
# CSS
# ==============================================================
st.markdown(f"""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Noto+Sans+JP:wght@300;400;500;700;900&family=DM+Mono:wght@400;500&display=swap');
  html, body, [class*="css"] {{ font-family: 'Noto Sans JP', 'メイリオ', sans-serif; }}

  .hero {{
    background: linear-gradient(135deg, {THEME["primary"]} 0%, {THEME["primary_lt"]} 60%, #3A7BD5 100%);
    border-radius: 16px; padding: 28px 32px; margin-bottom: 28px; color: white;
    position: relative; overflow: hidden;
  }}
  .hero::before {{
    content: "🚌"; position: absolute; right: 24px; top: 50%;
    transform: translateY(-50%); font-size: 80px; opacity: 0.12;
  }}
  .hero h1  {{ font-size: 24px; margin: 0; font-weight: 900; letter-spacing: 0.03em; }}
  .hero p   {{ font-size: 12px; margin: 8px 0 0; opacity: 0.75; line-height: 1.7; }}
  .hero .v-badge {{
    display: inline-block; background: rgba(255,255,255,0.25);
    border: 1px solid rgba(255,255,255,0.4); border-radius: 20px;
    padding: 2px 12px; font-size: 11px; font-weight: 700;
    letter-spacing: 0.08em; margin-bottom: 10px;
  }}

  .step-header {{ display: flex; align-items: center; gap: 12px; margin: 28px 0 16px; }}
  .step-num {{
    width: 36px; height: 36px; background: {THEME["primary"]}; color: white;
    border-radius: 50%; display: flex; align-items: center; justify-content: center;
    font-size: 15px; font-weight: 900; flex-shrink: 0; font-family: 'DM Mono', monospace;
  }}
  .step-title {{ font-size: 17px; font-weight: 700; color: {THEME["text"]}; }}
  .step-sub   {{ font-size: 12px; color: {THEME["text2"]}; margin-top: 2px; }}

  .metric-grid {{
    display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin: 16px 0;
  }}
  .metric-item {{
    background: {THEME["surface"]}; border: 1px solid {THEME["border"]};
    border-radius: 10px; padding: 16px 12px; text-align: center;
  }}
  .metric-item .val   {{ font-size: 32px; font-weight: 900; color: {THEME["primary"]}; line-height: 1; }}
  .metric-item .unit  {{ font-size: 13px; font-weight: 600; color: {THEME["primary"]}; }}
  .metric-item .label {{ font-size: 11px; color: {THEME["text2"]}; margin-top: 4px; }}

  .ok   {{ color: #27AE60; font-weight: 700; }}
  .fail {{ color: #E74C3C; font-weight: 700; }}

  .cal-date-badge {{
    background: {THEME["accent"]}; color: white; border-radius: 8px;
    padding: 4px 14px; font-size: 14px; font-weight: 700;
    font-family: 'DM Mono', monospace; display: inline-block; margin-bottom: 8px;
  }}

  .timeline-item {{
    display: flex; align-items: flex-start; gap: 12px; padding: 10px 0;
    border-bottom: 1px dashed {THEME["border"]};
  }}
  .timeline-item:last-child {{ border-bottom: none; }}
  .timeline-dot  {{ width: 10px; height: 10px; border-radius: 50%; background: {THEME["accent"]}; margin-top: 5px; flex-shrink: 0; }}
  .timeline-time {{ font-family: 'DM Mono', monospace; font-size: 13px; font-weight: 600; color: {THEME["accent"]}; width: 50px; flex-shrink: 0; }}
  .timeline-name {{ font-size: 13px; font-weight: 600; }}
  .timeline-detail {{ font-size: 11px; color: {THEME["text2"]}; }}

  .sidebar-section {{
    background: white; border-radius: 10px; padding: 14px 16px;
    margin-bottom: 12px; border: 1px solid {THEME["border"]};
  }}
  .sidebar-section h4 {{
    font-size: 13px; font-weight: 700; color: {THEME["primary"]};
    margin: 0 0 10px; padding-bottom: 6px; border-bottom: 1px solid {THEME["border"]};
  }}

  .stButton > button[kind="primary"] {{
    background: linear-gradient(135deg, {THEME["primary"]} 0%, {THEME["primary_lt"]} 100%) !important;
    color: white !important; border: none !important; border-radius: 10px !important;
    font-weight: 700 !important; font-size: 15px !important; padding: 14px 32px !important;
  }}
  .stButton > button[kind="secondary"] {{ border-radius: 8px !important; font-weight: 600 !important; }}
  .stDownloadButton > button {{
    border-radius: 10px !important; font-weight: 700 !important;
    background: {THEME["accent"]} !important; color: white !important; border: none !important;
  }}

  @keyframes fadeUp {{ from {{ opacity:0; transform:translateY(10px); }} to {{ opacity:1; transform:translateY(0); }} }}
  .fade-up {{ animation: fadeUp 0.3s ease; }}

  @media print {{
    header, section[data-testid="stSidebar"], .stButton, .stDownloadButton {{ display: none !important; }}
  }}
</style>
""", unsafe_allow_html=True)


# ==============================================================
# ユーティリティ関数
# ==============================================================

def hhmm_to_min(s, default: int = 0) -> int:
    """HH:MM 文字列 → 分数。変換失敗時は default を返す。"""
    if not s or str(s).strip() in ("", "nan", "None"):
        return default
    s = str(s).strip()
    try:
        parts = s.split(":")
        return int(parts[0]) * 60 + int(parts[1])
    except (ValueError, IndexError):
        try:
            return int(float(s))
        except (ValueError, TypeError):
            return default


def min_to_hhmm(m: int) -> str:
    """分数 → HH:MM 文字列。"""
    h, mn = divmod(abs(m), 60)
    return f"{h:02d}:{mn:02d}"


def step_header(num: int, title: str, sub: str = ""):
    st.markdown(f"""
    <div class="step-header fade-up">
      <div class="step-num">{num}</div>
      <div>
        <div class="step-title">{title}</div>
        {"<div class='step-sub'>" + sub + "</div>" if sub else ""}
      </div>
    </div>
    """, unsafe_allow_html=True)


def metric_row(items: list):
    cols = "".join([
        f'<div class="metric-item"><div class="val">{v}</div>'
        f'<div class="unit">{u}</div><div class="label">{l}</div></div>'
        for v, u, l in items
    ])
    st.markdown(f'<div class="metric-grid">{cols}</div>', unsafe_allow_html=True)


# ==============================================================
# データクラス
# ==============================================================

class ServiceType(Enum):
    HOUKAGO_DEI = "放課後等デイサービス"
    A_TYPE      = "就労継続支援A型"
    B_TYPE      = "就労継続支援B型"


class TripType(Enum):
    PICKUP  = "迎え"
    DROPOFF = "送り"


@dataclass
class User:
    user_id:          str
    name:             str
    address:          str
    lat:              float
    lng:              float
    service_type:     ServiceType
    shop:             str
    wheelchair:       bool = False
    incompatible:     list = field(default_factory=list)
    # ── v7 時間枠（分単位）──
    # 迎え便
    pickup_earliest:  int = 0     # 早迎え禁止時刻: Node CumulVar.SetMin。0=制約なし
    pickup_latest:    int = 540   # 迎えリミット:   Node CumulVar.SetMax
    # 送り便
    dropoff_earliest: int = 0     # 送り出発可能時刻: Node CumulVar.SetMin。0=制約なし
    dropoff_latest:   int = 0     # 送り到着リミット: Node CumulVar.SetMax。0=制約なし
    # 乗降時間（秒）
    service_time:     int = 300   # 通常5分。車椅子は10分。

    def __post_init__(self):
        if self.wheelchair:
            self.service_time = 600


@dataclass
class Vehicle:
    vehicle_id:    str
    name:          str
    vehicle_type:  str
    capacity:      int
    shop:          str
    wheelchair_ok: bool  = False
    depot_lat:     float = 36.6953
    depot_lng:     float = 137.2113


@dataclass
class Staff:
    staff_id:    str
    name:        str
    shop:        str
    can_drive:   bool = True
    priority:    int  = 1
    shift_start: Optional[int] = None   # 出勤時間（分）
    shift_end:   Optional[int] = None   # 退勤時間（分）


@dataclass
class AssignedRoute:
    vehicle:   Vehicle
    driver:    Optional[Staff]
    trip_type: TripType
    shop:      str
    stops:     list
    total_min: int


# ==============================================================
# 固定枠・列定義定数
# ==============================================================

SHOP_LIST = ["A店", "B店", "C店"]
STAFF_FRAMES: dict = {"A店": 10, "B店": 6, "C店": 6}
USER_FRAMES:  dict = {"A店": 30, "B店": 30, "C店": 30}

# 優先度 → Fixed Cost
PRIORITY_COST_MAP = {1: 0, 2: 5_000, 3: 15_000, 4: 30_000, 9: 999_999}

# スタッフマスタ列
STAFF_COL_NAME      = "B"
STAFF_COL_SHOP      = "C"
STAFF_COL_SHIFT_ST  = "F"
STAFF_COL_SHIFT_EN  = "G"

# 利用者マスタ列 (v7: A〜T)
# A=ID, B=氏名, C=住所, D=緯度, E=経度, F=サービス種別, G=店舗,
# H=車椅子, I=同乗不可ID,
# J=早迎え禁止時刻, K=迎えリミット, L=送り出発可能時刻, M=送り到着リミット,
# N=月, O=火, P=水, Q=木, R=金, S=土, T=日
USER_COL_NAME       = "B"
USER_COL_EARLY_PU   = "J"   # 早迎え禁止時刻
USER_COL_LIMIT_PU   = "K"   # 迎えリミット
USER_COL_EARLY_DO   = "L"   # 送り出発可能時刻
USER_COL_LIMIT_DO   = "M"   # 送り到着リミット

# 曜日列: 月=N(0), 火=O(1), 水=P(2), 木=Q(3), 金=R(4), 土=S(5), 日=T(6)
DOW_COLS = ["N", "O", "P", "Q", "R", "S", "T"]
DOW_NAMES = ["月", "火", "水", "木", "金", "土", "日"]


def _get_master_row_ranges(frames: dict) -> dict:
    """マスタシートの各店舗の行範囲 (1-indexed) を返す。Row1=タイトル, Row2=ヘッダー, Row3〜=データ。"""
    ranges = {}
    cur = 3
    for shop in SHOP_LIST:
        n = frames.get(shop, 0)
        ranges[shop] = (cur, cur + n - 1)
        cur += n
    return ranges


def _get_calendar_shop_layout(frames: dict) -> list:
    """カレンダーの各店舗ブロック行情報 [(shop, hdr_row, data_start, data_end, mst_start), ...]"""
    master_ranges = _get_master_row_ranges(frames)
    layout = []
    cal_row = 5   # Row1=タイトル, Row2=凡例, Row3=日付, Row4=曜日, Row5=最初の店舗ヘッダー
    for shop in SHOP_LIST:
        n = frames[shop]
        hdr_row    = cal_row
        data_start = cal_row + 1
        data_end   = data_start + n - 1
        mst_start  = master_ranges[shop][0]
        layout.append((shop, hdr_row, data_start, data_end, mst_start))
        cal_row = data_end + 2   # +1=空白行, +1=次ヘッダー
    return layout


# ==============================================================
# 距離行列ビルダー（ハーバーサイン距離）
# ==============================================================

class DistanceMatrixBuilder:
    def build(self, locations: list) -> list:
        n = len(locations)
        return [[self._h(locations[i], locations[j]) for j in range(n)] for i in range(n)]

    @staticmethod
    def _h(p1, p2, spd: float = 30.0) -> int:
        R = 6371.0
        la1, lo1 = math.radians(p1[0]), math.radians(p1[1])
        la2, lo2 = math.radians(p2[0]), math.radians(p2[1])
        dla, dlo = la2 - la1, lo2 - lo1
        a = math.sin(dla/2)**2 + math.cos(la1)*math.cos(la2)*math.sin(dlo/2)**2
        return max(1, int(2 * R * math.asin(math.sqrt(a)) / spd * 3600))


# ==============================================================
# 制約チェッカー
# ==============================================================

class ConstraintChecker:
    @staticmethod
    def validate(users, vehicles, staff) -> list:
        errors = []
        if not [s for s in staff if s.can_drive]:
            errors.append("❌ 運転可能なスタッフが0人です")
        wcu = [u for u in users if u.wheelchair]
        wcv = [v for v in vehicles if v.wheelchair_ok]
        if wcu and not wcv:
            errors.append(f"❌ 車椅子利用者 {[u.name for u in wcu]} がいますが車椅子対応車両がありません")
        if len(users) > sum(v.capacity for v in vehicles):
            errors.append(f"❌ 利用者数({len(users)}名) > 全車両定員合計({sum(v.capacity for v in vehicles)}名)")
        return errors

    @staticmethod
    def get_forbidden_pairs(users) -> set:
        pairs = set()
        for u in users:
            for iid in u.incompatible:
                pairs.add(tuple(sorted([u.user_id, iid])))
        return pairs


# ==============================================================
# VRP ソルバー（v7: 4項目の時間枠 + 乗降時間 + 優先度コスト）
# ==============================================================

class TransportVRPSolver:
    TIME_LIMIT_SEC = 30

    def __init__(self, users, vehicles, staff, distance_matrix,
                 trip_type=TripType.PICKUP,
                 depot_arrival_limit_min: int = 600,
                 start_time_min: int = 480):
        self.users     = users
        self.vehicles  = vehicles
        self.staff     = sorted([s for s in staff if s.can_drive], key=lambda s: s.priority)
        self.matrix    = distance_matrix
        self.trip_type = trip_type
        self.global_limit = depot_arrival_limit_min * 60
        self.start_time   = start_time_min * 60
        checker = ConstraintChecker()
        self.forbidden_pairs = checker.get_forbidden_pairs(users)

    def solve(self) -> list:
        if ORTOOLS_AVAILABLE and self.users:
            result = self._solve_ortools()
            if result is not None:
                return result
        return self._greedy()

    def _solve_ortools(self) -> Optional[list]:
        nu, nv = len(self.users), len(self.vehicles)
        if nv == 0 or nu == 0:
            return []
        nn = nu + 1

        manager = pywrapcp.RoutingIndexManager(nn, nv, 0)
        routing = pywrapcp.RoutingModel(manager)

        # ── コールバック: 移動時間 + 乗降時間 ──
        def time_callback(fi, ti):
            fn = manager.IndexToNode(fi)
            tn = manager.IndexToNode(ti)
            travel = self.matrix[fn][tn]
            svc    = self.users[fn - 1].service_time if fn > 0 else 0
            return travel + svc

        transit_cb = routing.RegisterTransitCallback(time_callback)
        routing.SetArcCostEvaluatorOfAllVehicles(transit_cb)

        # ── 時間ディメンション ──
        max_tw = max(
            self.global_limit,
            *([u.pickup_latest  * 60 for u in self.users if u.pickup_latest  > 0] or [0]),
            *([u.dropoff_latest * 60 for u in self.users if u.dropoff_latest > 0] or [0]),
            self.start_time + 8 * 3600,
        )
        routing.AddDimension(transit_cb, 600, max_tw, True, "Time")
        time_dim = routing.GetDimensionOrDie("Time")

        # ── v7: ユーザーノードごとの時間枠（シンプル4項目）──
        for i, user in enumerate(self.users):
            node_idx = manager.NodeToIndex(i + 1)

            if self.trip_type == TripType.PICKUP:
                # 早迎え禁止時刻: SetMin（0=制約なし）
                if user.pickup_earliest > 0:
                    earliest_sec = (user.pickup_earliest * 60) - self.start_time
                    if earliest_sec > 0:
                        time_dim.CumulVar(node_idx).SetMin(earliest_sec)
                # 迎えリミット: SetMax
                if user.pickup_latest > 0:
                    limit_sec = (user.pickup_latest * 60) - self.start_time
                    if limit_sec > 0:
                        time_dim.CumulVar(node_idx).SetMax(limit_sec)
            else:
                # 送り出発可能時刻: SetMin（0=制約なし）
                if user.dropoff_earliest > 0:
                    earl_sec = (user.dropoff_earliest * 60) - self.start_time
                    if earl_sec > 0:
                        time_dim.CumulVar(node_idx).SetMin(earl_sec)
                # 送り到着リミット: SetMax（0=制約なし）
                if user.dropoff_latest > 0:
                    dl_sec = (user.dropoff_latest * 60) - self.start_time
                    if dl_sec > 0:
                        time_dim.CumulVar(node_idx).SetMax(dl_sec)

        # デポの全体制限
        depot_idx = manager.NodeToIndex(0)
        time_dim.CumulVar(depot_idx).SetMax(self.global_limit)

        # ── シフト制約（スタッフ稼働時間）──
        for vi in range(nv):
            driver = self._get_driver(vi)
            if driver is None:
                continue
            if driver.shift_start is not None:
                ss_rel = max(0, (driver.shift_start * 60) - self.start_time)
                time_dim.CumulVar(routing.Start(vi)).SetMin(ss_rel)
            if driver.shift_end is not None:
                se_rel = (driver.shift_end * 60) - self.start_time
                if se_rel > 0:
                    time_dim.CumulVar(routing.End(vi)).SetMax(se_rel)

        # ── 定員制約 ──
        def demand_cb(fi):
            return 0 if manager.IndexToNode(fi) == 0 else 1

        demand_idx = routing.RegisterUnaryTransitCallback(demand_cb)
        routing.AddDimensionWithVehicleCapacity(
            demand_idx, 0, [v.capacity for v in self.vehicles], True, "Capacity"
        )

        # ── Fixed Cost: 優先度が低いほど高コスト ──
        for vi in range(nv):
            driver = self._get_driver(vi)
            p = driver.priority if driver else 9
            routing.SetFixedCostOfVehicle(PRIORITY_COST_MAP.get(p, 999_999), vi)

        # ── 車椅子制約 ──
        for i, u in enumerate(self.users):
            if u.wheelchair:
                ni = manager.NodeToIndex(i + 1)
                for vi, v in enumerate(self.vehicles):
                    if not v.wheelchair_ok:
                        routing.VehicleVar(ni).RemoveValue(vi)

        # ── 同乗不可制約 ──
        for uid1, uid2 in self.forbidden_pairs:
            i1 = next((i+1 for i, u in enumerate(self.users) if u.user_id == uid1), None)
            i2 = next((i+1 for i, u in enumerate(self.users) if u.user_id == uid2), None)
            if i1 and i2:
                ni1 = manager.NodeToIndex(i1)
                ni2 = manager.NodeToIndex(i2)
                routing.solver().Add(routing.VehicleVar(ni1) != routing.VehicleVar(ni2))

        # ── 探索パラメータ ──
        params = pywrapcp.DefaultRoutingSearchParameters()
        params.first_solution_strategy = routing_enums_pb2.FirstSolutionStrategy.PATH_CHEAPEST_ARC
        params.local_search_metaheuristic = routing_enums_pb2.LocalSearchMetaheuristic.GUIDED_LOCAL_SEARCH
        params.time_limit.FromSeconds(self.TIME_LIMIT_SEC)

        solution = routing.SolveWithParameters(params)
        if not solution:
            return None

        time_dim = routing.GetDimensionOrDie("Time")
        routes = []
        for vi in range(nv):
            idx   = routing.Start(vi)
            stops = []
            while not routing.IsEnd(idx):
                node = manager.IndexToNode(idx)
                if node != 0:
                    u       = self.users[node - 1]
                    arr_sec = solution.Min(time_dim.CumulVar(idx))
                    arr_min = (self.start_time + arr_sec) // 60
                    stops.append({"user": u, "arrival_min": arr_min,
                                  "address": u.address, "lat": u.lat, "lng": u.lng})
                idx = solution.Value(routing.NextVar(idx))
            if not stops:
                continue
            driver = self._get_driver(vi)
            routes.append(AssignedRoute(
                vehicle=self.vehicles[vi], driver=driver,
                trip_type=self.trip_type, shop=self.vehicles[vi].shop,
                stops=stops, total_min=0,
            ))
        return routes

    def _greedy(self) -> list:
        vs = sorted(self.vehicles, key=lambda v: (0 if v.wheelchair_ok else 1, -v.capacity))
        routes, unassigned = [], list(self.users)

        for vi, vehicle in enumerate(vs):
            if not unassigned:
                break
            driver   = self._get_driver(vi)
            assigned = []

            if vehicle.wheelchair_ok:
                for u in [u for u in unassigned if u.wheelchair]:
                    if len(assigned) < vehicle.capacity:
                        assigned.append(u); unassigned.remove(u)

            fids = set()
            for a in assigned:
                fids.update(a.incompatible)
            for u in list(unassigned):
                if len(assigned) >= vehicle.capacity: break
                if u.wheelchair and not vehicle.wheelchair_ok: continue
                if u.user_id in fids: continue
                assigned.append(u); unassigned.remove(u); fids.update(u.incompatible)

            if not assigned:
                continue

            ordered  = self._nn(assigned)
            stops    = []
            cur_node = 0
            cur_time = self.start_time

            for u in ordered:
                uid       = self.users.index(u) + 1
                cur_time += self.matrix[cur_node][uid] + u.service_time
                stops.append({"user": u, "arrival_min": cur_time // 60,
                              "address": u.address, "lat": u.lat, "lng": u.lng})
                cur_node = uid

            routes.append(AssignedRoute(
                vehicle=vehicle, driver=driver,
                trip_type=self.trip_type, shop=vehicle.shop,
                stops=stops, total_min=(cur_time - self.start_time) // 60,
            ))

        if unassigned:
            st.warning(f"⚠️ [{self.vehicles[0].shop if self.vehicles else ''}] 未割当: {[u.name for u in unassigned]}")
        return routes

    def _nn(self, users: list) -> list:
        if not users:
            return []
        rem, ordered, cur = list(users), [], 0
        while rem:
            n = min(rem, key=lambda u: self.matrix[cur][self.users.index(u) + 1])
            ordered.append(n); cur = self.users.index(n) + 1; rem.remove(n)
        return ordered

    def _get_driver(self, vi: int) -> Optional[Staff]:
        return self.staff[vi % len(self.staff)] if self.staff else None


# ==============================================================
# 店舗別VRP実行（シフトフィルタリング付き）
# ==============================================================

def run_all_shops(users, vehicles, staff, trip_type, start_min, limit_min) -> list:
    """店舗ごとにデータを分割し、独立したVRPを実行してマージして返す。"""
    shops  = sorted(set(u.shop for u in users))
    routes = []

    for shop in shops:
        su = [u for u in users    if u.shop == shop]
        sv = [v for v in vehicles if v.shop == shop]

        def is_on_shift(s: Staff) -> bool:
            if not s.can_drive:
                return False
            if s.shift_start is not None and s.shift_end is not None:
                return s.shift_start <= start_min < s.shift_end
            if s.shift_start is not None:
                return s.shift_start <= start_min
            if s.shift_end is not None:
                return start_min < s.shift_end
            return True

        ss = [s for s in staff if s.shop == shop and is_on_shift(s)]

        if not su or not sv:
            continue
        if not ss:
            st.warning(f"⚠️ [{shop}] 該当時間帯に勤務可能なスタッフがいません")
            continue

        depot  = (sv[0].depot_lat, sv[0].depot_lng)
        locs   = [depot] + [(u.lat, u.lng) for u in su]
        matrix = DistanceMatrixBuilder().build(locs)

        solver = TransportVRPSolver(
            users=su, vehicles=sv, staff=ss, distance_matrix=matrix,
            trip_type=trip_type, depot_arrival_limit_min=limit_min, start_time_min=start_min,
        )
        routes.extend(solver.solve())

    return routes


# ==============================================================
# カレンダーセル値パーサー（堅牢版）
# ==============================================================

def parse_time_range(cell_val, default_start: int = 480, default_end: int = 1140):
    """
    カレンダーセルの値を解析して (start_min, end_min) または None を返す。

    対応フォーマット:
      "HH:MM-HH:MM" → (start, end)
      "HH:MM"       → (start, default_end)
      "〇" / "○"    → (default_start, default_end)
      "00:00-HH:MM" → (0, end)  ← 早迎え禁止なし（0=制約なし）
      空欄 / nan    → None（欠席・休み）
      "="始まり     → None（未計算数式）
    """
    if cell_val is None:
        return None
    s = str(cell_val).strip()
    if s in ("", "nan", "None", "-", "×", "✕", "欠席", "休み", "休"):
        return None
    if s.startswith("="):
        return None  # 未計算数式
    if s in ("〇", "○", "◯", "✓", "✔", "出", "◎"):
        return (default_start, default_end)

    # "HH:MM-HH:MM" 形式（各種ハイフン対応）
    for sep in ["-", "～", "〜", "~", "ー", "−", "–"]:
        if sep in s:
            parts = s.split(sep, 1)
            if len(parts) == 2:
                try:
                    st_ = hhmm_to_min(parts[0].strip(), -1)
                    en  = hhmm_to_min(parts[1].strip(), -1)
                    if st_ >= 0 and en > 0:
                        return (st_, en)
                except Exception:
                    pass

    # "HH:MM" 単体
    try:
        st_ = hhmm_to_min(s, -1)
        if st_ >= 0:
            return (st_, default_end)
    except Exception:
        pass

    return None


def _is_empty_cell(val) -> bool:
    """空セル・NaN・未計算数式 → True"""
    if val is None:
        return True
    s = str(val).strip()
    return s in ("", "nan", "None") or s.startswith("=")


# ==============================================================
# Excel 入力: parse_excel_upload
# ==============================================================

def parse_excel_upload(
    uploaded_file,
    default_pickup_limit: int = 540,
    default_dropoff_limit: int = 1110,
):
    """
    Excelを読み込んでデータクラスに変換。
    v7: 新しい4項目の時間カラム + 曜日カラム（読み込み不要、数式参照用）
    空行・数式文字列は自動スキップ。
    """
    xl = pd.ExcelFile(uploaded_file)
    svc_map = {
        "放課後等デイサービス": ServiceType.HOUKAGO_DEI,
        "A型": ServiceType.A_TYPE,
        "B型": ServiceType.B_TYPE,
    }

    # ---- 利用者 ----
    df_u = xl.parse("利用者", header=1)
    users = []
    for i, row in df_u.iterrows():
        name_val = row.get("氏名", "")
        if _is_empty_cell(name_val):
            continue  # 空行スキップ

        incomp_raw = str(row.get("同乗不可ID", "")).strip()
        incomp = ([x.strip() for x in incomp_raw.split(",") if x.strip()]
                  if incomp_raw not in ("", "nan") else [])
        wc = bool(row.get("車椅子", False))

        def _rt(key, default):
            v = row.get(key, "")
            if _is_empty_cell(v): return default
            return hhmm_to_min(str(v).strip(), default)

        # v7 4項目 + 後方互換（旧カラム名も受け付ける）
        pu_earl = _rt("早迎え禁止時刻", 0)
        pu_last = _rt("迎えリミット",   _rt("到着リミット", default_pickup_limit))
        do_earl = _rt("送り出発可能時刻", 0)
        do_last = _rt("送り到着リミット", _rt("送り目標", default_dropoff_limit) if _rt("送り目標", 0) else 0)

        try:
            lat = float(row.get("緯度", 36.695))
        except (ValueError, TypeError):
            lat = 36.695
        try:
            lng = float(row.get("経度", 137.211))
        except (ValueError, TypeError):
            lng = 137.211

        users.append(User(
            user_id          = str(row.get("ID", f"u{i+1}")),
            name             = str(name_val).strip(),
            address          = str(row.get("住所", "")),
            lat              = lat,
            lng              = lng,
            service_type     = svc_map.get(str(row.get("サービス種別", "")), ServiceType.HOUKAGO_DEI),
            shop             = str(row.get("店舗", "A店")),
            wheelchair       = wc,
            incompatible     = incomp,
            pickup_earliest  = pu_earl,
            pickup_latest    = pu_last,
            dropoff_earliest = do_earl,
            dropoff_latest   = do_last,
            service_time     = 600 if wc else 300,
        ))

    # ---- 車両 ----
    df_v    = xl.parse("車両", header=1)
    type_cap = {"large": 7, "normal": 4, "kei": 3}
    vehicles = []
    for i, row in df_v.iterrows():
        vname = row.get("車両名", "")
        if _is_empty_cell(vname):
            continue
        vtype = str(row.get("種別コード", "normal"))
        try:
            dlat = float(row.get("デポ緯度", 36.695))
        except (ValueError, TypeError):
            dlat = 36.695
        try:
            dlng = float(row.get("デポ経度", 137.211))
        except (ValueError, TypeError):
            dlng = 137.211
        vehicles.append(Vehicle(
            vehicle_id    = str(row.get("ID", f"v{i+1}")),
            name          = str(vname).strip(),
            vehicle_type  = vtype,
            capacity      = int(row.get("定員", type_cap.get(vtype, 4))),
            shop          = str(row.get("店舗", "A店")),
            wheelchair_ok = bool(row.get("車椅子対応", False)),
            depot_lat     = dlat,
            depot_lng     = dlng,
        ))

    # ---- スタッフ ----
    df_s  = xl.parse("スタッフ", header=1)
    staff = []
    for i, row in df_s.iterrows():
        name_val = row.get("氏名", "")
        if _is_empty_cell(name_val):
            continue
        ss_raw = row.get("出勤時間", "")
        se_raw = row.get("退勤時間", "")
        ss = hhmm_to_min(ss_raw, -1) if not _is_empty_cell(ss_raw) else None
        se = hhmm_to_min(se_raw, -1) if not _is_empty_cell(se_raw) else None
        staff.append(Staff(
            staff_id    = str(row.get("ID", f"s{i+1}")),
            name        = str(name_val).strip(),
            shop        = str(row.get("店舗", "A店")),
            can_drive   = bool(row.get("運転可否", True)),
            priority    = int(row.get("優先度", 1)),
            shift_start = ss if ss != -1 else None,
            shift_end   = se if se != -1 else None,
        ))

    # ---- カレンダー ----
    calendar_data = _parse_calendar_sheets(xl, staff, users)

    return users, vehicles, staff, calendar_data


# ==============================================================
# カレンダーパース
# ==============================================================

def _parse_calendar_sheet(df, name_col, first_day_col, year, month,
                           default_start=480, default_end=1140) -> dict:
    """横軸=日付、縦軸=名前 のカレンダーシートを解析。"""
    _, days_in_month = cal_mod.monthrange(year, month)
    result = {}

    SKIP_KW = ["🏠", "📅", "📋", "【入力", "【運用", "スタッフ シフト",
               "利用者 月間", "スタッフ名", "利用者氏名", "氏名"]

    for row_idx in range(len(df)):
        row      = df.iloc[row_idx]
        name_val = row.iloc[name_col] if name_col < len(row) else None

        if name_val is None or str(name_val).strip() in ("", "nan"):
            continue
        name = str(name_val).strip()
        if name.startswith("="):
            continue
        if any(kw in name for kw in SKIP_KW):
            continue
        name = name.replace("♿ ", "").replace("♿", "").strip()

        schedule = {}
        for day in range(1, days_in_month + 1):
            day_col_idx = first_day_col + (day - 1)
            if day_col_idx >= len(row):
                break
            cell_val = row.iloc[day_col_idx]
            date_str = datetime.date(year, month, day).strftime("%Y-%m-%d")
            if cell_val is not None and str(cell_val).strip().startswith("="):
                schedule[date_str] = None  # 未計算数式 = 欠席扱い
            else:
                schedule[date_str] = parse_time_range(cell_val, default_start, default_end)

        result[name] = schedule

    return result


def _detect_year_month_from_sheet(df_raw, xl, sheet_name) -> tuple:
    try:
        title_val = str(df_raw.iloc[0, 0])
        m = re.search(r"(\d{4})年\s*(\d{1,2})月", title_val)
        if m:
            return int(m.group(1)), int(m.group(2))
    except Exception:
        pass
    today = datetime.date.today()
    return today.year, today.month


def _parse_calendar_sheets(xl, staff, users) -> Optional[dict]:
    has_s = "カレンダー_スタッフ" in xl.sheet_names
    has_u = "カレンダー_利用者"   in xl.sheet_names
    if not has_s and not has_u:
        return None

    today = datetime.date.today()
    year, month = today.year, today.month
    staff_cal, users_cal = {}, {}

    if has_s:
        raw_s = xl.parse("カレンダー_スタッフ", header=None)
        year, month = _detect_year_month_from_sheet(raw_s, xl, "カレンダー_スタッフ")
        staff_cal   = _parse_calendar_sheet(raw_s, 0, 2, year, month, 480, 1140)

    if has_u:
        raw_u = xl.parse("カレンダー_利用者", header=None)
        year, month = _detect_year_month_from_sheet(raw_u, xl, "カレンダー_利用者")
        users_cal   = _parse_calendar_sheet(raw_u, 0, 2, year, month, 480, 540)

    return {"staff": staff_cal, "users": users_cal, "year": year, "month": month}


# ==============================================================
# Excel 書き込み: マスタシート（固定枠・店舗ブロック付き）
# ==============================================================

TIME_COLS_SET = {
    "出勤時間", "退勤時間",
    "早迎え禁止時刻", "迎えリミット", "送り出発可能時刻", "送り到着リミット",
    "到着リミット", "送り目標",
}


def _cell_bdr(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _write_master_sheet_v5(ws, shop_data, frames, headers, title,
                            header_color="2C4A6E", shop_col_key="店舗"):
    """固定枠付き・店舗ブロック形式のマスタシート書き込み。"""
    TITLE_FILL = PatternFill("solid", fgColor="1B3A5C")
    HDR_FILL   = PatternFill("solid", fgColor=header_color)
    SHOP_FILLS = {
        "A店": PatternFill("solid", fgColor="D4E6F1"),
        "B店": PatternFill("solid", fgColor="D5F5E3"),
        "C店": PatternFill("solid", fgColor="FEF9E7"),
    }
    EMPTY_FILL = PatternFill("solid", fgColor="FAFBFC")
    ODD_FILL   = PatternFill("solid", fgColor="FFFFFF")

    n_cols = len(headers)
    W = get_column_letter

    # タイトル行
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font      = Font(bold=True, size=12, color="FFFFFF", name="メイリオ")
    c.fill      = TITLE_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ヘッダー行
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=10, name="メイリオ")
        c.fill      = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _cell_bdr()
    ws.row_dimensions[2].height = 22

    master_ranges = _get_master_row_ranges(frames)
    for shop in SHOP_LIST:
        start_row, _ = master_ranges[shop]
        data_rows    = shop_data.get(shop, [])
        shop_fill    = SHOP_FILLS.get(shop, EMPTY_FILL)

        for slot_idx in range(frames[shop]):
            row_idx  = start_row + slot_idx
            is_empty = slot_idx >= len(data_rows)
            row_data = data_rows[slot_idx] if not is_empty else {h: "" for h in headers}
            if is_empty:
                row_data[shop_col_key] = shop

            fill = EMPTY_FILL if is_empty else (shop_fill if slot_idx % 2 == 0 else ODD_FILL)

            for col, h in enumerate(headers, 1):
                val = row_data.get(h, "")
                c   = ws.cell(row=row_idx, column=col, value=val)
                c.font      = Font(
                    size=10, name="メイリオ",
                    color="AAAAAA" if is_empty and h not in (shop_col_key, "ID") else "1C2330"
                )
                c.fill      = fill
                c.alignment = Alignment(
                    horizontal="left" if isinstance(val, str) and len(str(val)) > 6 else "center",
                    vertical="center"
                )
                c.border = _cell_bdr()
                if h in TIME_COLS_SET:
                    c.number_format = "@"  # Text形式でシリアル値変換を防止
                # 曜日列は中央寄せ
                if h in DOW_NAMES:
                    c.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[row_idx].height = 20

    # 列幅調整
    all_rows = [r for rows in shop_data.values() for r in rows]
    for col, h in enumerate(headers, 1):
        max_len = max(
            len(str(h)),
            max((len(str(r.get(h, ""))) for r in all_rows), default=0)
        )
        ws.column_dimensions[W(col)].width = min(max(max_len + 2, 6), 30)
    # 曜日列は狭く
    for dow in DOW_NAMES:
        if dow in headers:
            col_idx = headers.index(dow) + 1
            ws.column_dimensions[W(col_idx)].width = 5


def _write_master_sheet(ws, rows, title, header_color="2C4A6E"):
    """固定枠なし版（車両マスタ等）。"""
    if not rows:
        return
    TITLE_FILL = PatternFill("solid", fgColor="1B3A5C")
    HDR_FILL   = PatternFill("solid", fgColor=header_color)
    EVEN_FILL  = PatternFill("solid", fgColor="F8F9FA")
    ODD_FILL   = PatternFill("solid", fgColor="FFFFFF")

    headers = list(rows[0].keys())
    n_cols  = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font      = Font(bold=True, size=12, color="FFFFFF", name="メイリオ")
    c.fill      = TITLE_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=10, name="メイリオ")
        c.fill      = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _cell_bdr()
        max_len = max(len(str(h)), max((len(str(r.get(h, ""))) for r in rows), default=0))
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 8), 40)
    ws.row_dimensions[2].height = 22

    for ri, row_data in enumerate(rows):
        row_idx = ri + 3
        fill    = EVEN_FILL if ri % 2 == 0 else ODD_FILL
        for col, h in enumerate(headers, 1):
            val = row_data.get(h, "")
            c   = ws.cell(row=row_idx, column=col, value=val)
            c.font      = Font(size=10, name="メイリオ")
            c.fill      = fill
            c.alignment = Alignment(
                horizontal="left" if isinstance(val, str) and len(str(val)) > 6 else "center",
                vertical="center"
            )
            c.border = _cell_bdr()
        ws.row_dimensions[row_idx].height = 20


# ==============================================================
# Excel 書き込み: スタッフカレンダー（数式埋め込み版）
# ==============================================================

def _write_calendar_sheet_staff(wb, staff: list, year: int, month: int):
    """
    スタッフシフト表シート。
    平日（日曜以外）: =IF(スタッフ!$B{n}="","",IF(スタッフ!$F{n}="","",
                         TEXT(IFERROR(TIMEVALUE(スタッフ!$F{n}),スタッフ!$F{n}),"HH:MM")
                         &"-"&
                         TEXT(IFERROR(TIMEVALUE(スタッフ!$G{n}),スタッフ!$G{n}),"HH:MM")))
    日曜: 空欄固定
    """
    _, days_in_month = cal_mod.monthrange(year, month)
    dates = [datetime.date(year, month, d) for d in range(1, days_in_month + 1)]

    ws = wb.create_sheet("カレンダー_スタッフ")

    C_TITLE   = PatternFill("solid", fgColor="1B3A5C")
    C_HDR     = PatternFill("solid", fgColor="2C4A6E")
    C_SAT     = PatternFill("solid", fgColor="EBF5FB")
    C_SUN     = PatternFill("solid", fgColor="FDECEA")
    C_FORMULA = PatternFill("solid", fgColor="F0FFF4")
    C_WEEKEND = PatternFill("solid", fgColor="F5F5F5")
    SHOP_LIGHT = {"A店": "D4E6F1", "B店": "D5F5E3", "C店": "FEF9E7"}
    SHOP_DARK  = {"A店": "1A5276", "B店": "145A32", "C店": "6E2F1A"}

    NAME_COL  = 1
    SHOP_COL  = 2
    FIRST_DAY = 3
    last_col  = FIRST_DAY + days_in_month - 1

    # Row1: タイトル
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(row=1, column=1, value=f"📅 スタッフ シフト表　{year}年{month}月")
    c.font      = Font(bold=True, size=14, color="FFFFFF", name="メイリオ")
    c.fill      = C_TITLE
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Row2: 凡例
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    c = ws.cell(row=2, column=1,
                value="【運用】氏名・時間はマスタから自動転記。変更がある日のセルを直接上書き入力。空欄=休み。")
    c.font  = Font(italic=True, size=10, color="555555", name="メイリオ")
    c.fill  = PatternFill("solid", fgColor="EAF2FA")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    # Row3〜4: 日付ヘッダー
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
    ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
    for col, lbl in [(1, "氏名"), (2, "店舗")]:
        c = ws.cell(row=3, column=col, value=lbl)
        c.font = Font(bold=True, color="FFFFFF", size=10, name="メイリオ")
        c.fill = C_HDR; c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _cell_bdr()

    DAY_NAMES = ["月", "火", "水", "木", "金", "土", "日"]
    for di, dt in enumerate(dates):
        col    = FIRST_DAY + di
        is_sat = dt.weekday() == 5
        is_sun = dt.weekday() == 6
        hfill  = C_SUN if is_sun else (C_SAT if is_sat else C_HDR)
        hcolor = "C0392B" if is_sun else ("1A5276" if is_sat else "FFFFFF")

        c = ws.cell(row=3, column=col, value=int(dt.day))
        c.font = Font(bold=True, size=10, color=hcolor, name="メイリオ")
        c.fill = hfill; c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _cell_bdr()

        c = ws.cell(row=4, column=col, value=DAY_NAMES[dt.weekday()])
        c.font = Font(bold=True, size=9, color=hcolor, name="メイリオ")
        c.fill = hfill; c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _cell_bdr()
        ws.column_dimensions[get_column_letter(col)].width = 10

    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 16
    ws.column_dimensions[get_column_letter(NAME_COL)].width = 14
    ws.column_dimensions[get_column_letter(SHOP_COL)].width = 7

    # データ行（店舗ブロック）
    layout = _get_calendar_shop_layout(STAFF_FRAMES)
    for shop, hdr_row, data_start, data_end, mst_start in layout:
        light = SHOP_LIGHT.get(shop, "F0F0F0")
        dark  = SHOP_DARK.get(shop, "2C3E50")
        n_slots = STAFF_FRAMES[shop]

        ws.merge_cells(start_row=hdr_row, start_column=1, end_row=hdr_row, end_column=last_col)
        c = ws.cell(row=hdr_row, column=1, value=f"  🏠 {shop}")
        c.font      = Font(bold=True, size=11, color="FFFFFF", name="メイリオ")
        c.fill      = PatternFill("solid", fgColor="2C3E50")
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _cell_bdr("2C3E50")
        ws.row_dimensions[hdr_row].height = 22

        for slot_idx in range(n_slots):
            cal_row = data_start + slot_idx
            mst_row = mst_start + slot_idx

            # 氏名数式
            name_f = f'=IF(スタッフ!${STAFF_COL_NAME}{mst_row}="","",スタッフ!${STAFF_COL_NAME}{mst_row})'
            c = ws.cell(row=cal_row, column=NAME_COL, value=name_f)
            c.font = Font(bold=True, size=10, name="メイリオ")
            c.fill = PatternFill("solid", fgColor=light)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = _cell_bdr()

            # 店舗数式
            shop_f = f'=IF(スタッフ!${STAFF_COL_NAME}{mst_row}="","",スタッフ!${STAFF_COL_SHOP}{mst_row})'
            c = ws.cell(row=cal_row, column=SHOP_COL, value=shop_f)
            c.font = Font(size=9, name="メイリオ")
            c.fill = PatternFill("solid", fgColor=light)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _cell_bdr()

            # 各日セル
            for di, dt in enumerate(dates):
                col    = FIRST_DAY + di
                is_sun = dt.weekday() == 6
                is_sat = dt.weekday() == 5

                if is_sun:
                    # 日曜: 空欄
                    c = ws.cell(row=cal_row, column=col, value="")
                    c.fill   = C_WEEKEND
                    c.border = _cell_bdr()
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    # 月〜土: マスタのシフト時間を参照
                    # TEXT(IFERROR(TIMEVALUE(...), ...)) でテキスト・シリアル値両対応
                    f_st = f'IFERROR(TEXT(TIMEVALUE(スタッフ!${STAFF_COL_SHIFT_ST}{mst_row}),"HH:MM"),スタッフ!${STAFF_COL_SHIFT_ST}{mst_row})'
                    f_en = f'IFERROR(TEXT(TIMEVALUE(スタッフ!${STAFF_COL_SHIFT_EN}{mst_row}),"HH:MM"),スタッフ!${STAFF_COL_SHIFT_EN}{mst_row})'
                    time_f = (
                        f'=IF(スタッフ!${STAFF_COL_NAME}{mst_row}="","",IF(スタッフ!${STAFF_COL_SHIFT_ST}{mst_row}="","",{f_st}&"-"&{f_en}))'
                    )
                    c = ws.cell(row=cal_row, column=col, value=time_f)
                    c.fill   = C_SAT if is_sat else C_FORMULA
                    c.border = _cell_bdr()
                    c.font   = Font(size=9, color="1E6B38", name="メイリオ")
                    c.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[cal_row].height = 22

        sep_row = data_end + 1
        for col in range(1, last_col + 1):
            ws.cell(row=sep_row, column=col).fill = PatternFill("solid", fgColor="F0F3F4")
        ws.row_dimensions[sep_row].height = 6

    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6)
    ws.freeze_panes = f"{get_column_letter(FIRST_DAY)}5"


# ==============================================================
# Excel 書き込み: 利用者カレンダー（曜日×マスタ参照数式版）
# ==============================================================

def _write_calendar_sheet_users(wb, users: list, year: int, month: int):
    """
    利用者予定表シート。
    各日セルの数式: その日の曜日列に"〇"があれば "早迎え禁止-迎えリミット" を表示。
    曜日列 N=月, O=火, P=水, Q=木, R=金, S=土, T=日
    """
    _, days_in_month = cal_mod.monthrange(year, month)
    dates = [datetime.date(year, month, d) for d in range(1, days_in_month + 1)]

    ws = wb.create_sheet("カレンダー_利用者")

    C_TITLE   = PatternFill("solid", fgColor="1B3A5C")
    C_HDR     = PatternFill("solid", fgColor="2C4A6E")
    C_SAT     = PatternFill("solid", fgColor="EBF5FB")
    C_SUN     = PatternFill("solid", fgColor="FDECEA")
    C_FORMULA = PatternFill("solid", fgColor="F0FFF4")
    C_WC      = PatternFill("solid", fgColor="FEF9E7")
    C_WEEKEND = PatternFill("solid", fgColor="F5F5F5")
    SHOP_DARK  = {"A店": "1A5276", "B店": "145A32", "C店": "6E2F1A"}
    SHOP_LIGHT = {"A店": "D4E6F1", "B店": "D5F5E3", "C店": "FDEBD0"}

    NAME_COL  = 1
    SVC_COL   = 2
    FIRST_DAY = 3
    last_col  = FIRST_DAY + days_in_month - 1
    DAY_NAMES = ["月", "火", "水", "木", "金", "土", "日"]

    # Row1: タイトル
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(row=1, column=1, value=f"📋 利用者 月間利用予定表　{year}年{month}月")
    c.font      = Font(bold=True, size=14, color="FFFFFF", name="メイリオ")
    c.fill      = C_TITLE
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Row2: 凡例
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    c = ws.cell(row=2, column=1,
                value="【運用】氏名・時間はマスタの曜日設定から自動転記。例外の日は直接上書き入力。空欄=欠席・利用なし。")
    c.font  = Font(italic=True, size=10, color="555555", name="メイリオ")
    c.fill  = PatternFill("solid", fgColor="E8F8F5")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    # Row3〜4: 日付ヘッダー
    ws.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
    ws.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
    for col, lbl in [(1, "利用者氏名"), (2, "サービス")]:
        c = ws.cell(row=3, column=col, value=lbl)
        c.font = Font(bold=True, color="FFFFFF", size=10, name="メイリオ")
        c.fill = C_HDR; c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _cell_bdr()

    for di, dt in enumerate(dates):
        col    = FIRST_DAY + di
        is_sat = dt.weekday() == 5
        is_sun = dt.weekday() == 6
        hfill  = C_SUN if is_sun else (C_SAT if is_sat else C_HDR)
        hcolor = "C0392B" if is_sun else ("1A5276" if is_sat else "FFFFFF")

        c = ws.cell(row=3, column=col, value=int(dt.day))
        c.font = Font(bold=True, size=10, color=hcolor, name="メイリオ")
        c.fill = hfill; c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _cell_bdr()

        c = ws.cell(row=4, column=col, value=DAY_NAMES[dt.weekday()])
        c.font = Font(bold=True, size=9, color=hcolor, name="メイリオ")
        c.fill = hfill; c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _cell_bdr()
        ws.column_dimensions[get_column_letter(col)].width = 10

    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 16
    ws.column_dimensions[get_column_letter(NAME_COL)].width = 14
    ws.column_dimensions[get_column_letter(SVC_COL)].width  = 10

    # データ行（店舗ブロック）
    layout = _get_calendar_shop_layout(USER_FRAMES)
    for shop, hdr_row, data_start, data_end, mst_start in layout:
        dark  = SHOP_DARK.get(shop, "2C3E50")
        light = SHOP_LIGHT.get(shop, "F0F0F0")
        n_slots = USER_FRAMES[shop]

        ws.merge_cells(start_row=hdr_row, start_column=1, end_row=hdr_row, end_column=last_col)
        c = ws.cell(row=hdr_row, column=1, value=f"  🏠 {shop}　（{n_slots}名枠）")
        c.font      = Font(bold=True, size=11, color="FFFFFF", name="メイリオ")
        c.fill      = PatternFill("solid", fgColor=dark)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _cell_bdr(dark)
        ws.row_dimensions[hdr_row].height = 24

        for slot_idx in range(n_slots):
            cal_row = data_start + slot_idx
            mst_row = mst_start + slot_idx

            # 氏名数式
            name_f = f'=IF(利用者!${USER_COL_NAME}{mst_row}="","",利用者!${USER_COL_NAME}{mst_row})'
            c = ws.cell(row=cal_row, column=NAME_COL, value=name_f)
            c.font = Font(bold=True, size=10, name="メイリオ")
            c.fill = PatternFill("solid", fgColor=light)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = _cell_bdr()

            # サービス数式
            svc_f = f'=IF(利用者!${USER_COL_NAME}{mst_row}="","",利用者!$F{mst_row})'
            c = ws.cell(row=cal_row, column=SVC_COL, value=svc_f)
            c.font = Font(size=9, name="メイリオ")
            c.fill = PatternFill("solid", fgColor=light)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _cell_bdr()

            # 各日セル
            for di, dt in enumerate(dates):
                col      = FIRST_DAY + di
                is_sat   = dt.weekday() == 5
                is_sun   = dt.weekday() == 6
                dow_col  = DOW_COLS[dt.weekday()]   # 曜日列（N〜T）

                # 早迎え禁止時刻と迎えリミットをIFERROR+TIMEVALUEで両方に対応
                # 早迎え禁止が空の場合は "00:00" を表示 → parse_time_range で 0 に変換
                f_early = (
                    f'IF(利用者!${USER_COL_EARLY_PU}{mst_row}="","00:00",'
                    f'IFERROR(TEXT(TIMEVALUE(利用者!${USER_COL_EARLY_PU}{mst_row}),"HH:MM"),'
                    f'利用者!${USER_COL_EARLY_PU}{mst_row}))'
                )
                f_limit = (
                    f'IFERROR(TEXT(TIMEVALUE(利用者!${USER_COL_LIMIT_PU}{mst_row}),"HH:MM"),'
                    f'利用者!${USER_COL_LIMIT_PU}{mst_row})'
                )
                time_f = (
                    f'=IF(利用者!${USER_COL_NAME}{mst_row}="","",IF(利用者!${dow_col}{mst_row}="〇",'
                    f'{f_early}&"-"&{f_limit},""))'
                )

                c = ws.cell(row=cal_row, column=col, value=time_f)
                c.fill   = C_SUN if is_sun else (C_SAT if is_sat else C_FORMULA)
                c.border = _cell_bdr()
                c.font   = Font(size=9, color="1E6B38", name="メイリオ")
                c.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[cal_row].height = 22

        sep_row = data_end + 1
        for col in range(1, last_col + 1):
            ws.cell(row=sep_row, column=col).fill = PatternFill("solid", fgColor="F0F3F4")
        ws.row_dimensions[sep_row].height = 8

    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.6, bottom=0.6)
    ws.freeze_panes = f"{get_column_letter(FIRST_DAY)}5"


# ==============================================================
# Excel 書き込み: 送迎ルート表（v3〜継続）
# ==============================================================

def _write_route_sheet(ws, routes, sheet_title, date_str=""):
    TITLE_FILL = PatternFill("solid", fgColor="1B3A5C")
    DATE_FILL  = PatternFill("solid", fgColor="2D5F8A")
    HDR_FILL   = PatternFill("solid", fgColor="2C4A6E")
    WC_FILL    = PatternFill("solid", fgColor="FADBD8")
    ALT_FILL   = PatternFill("solid", fgColor="F6F8FA")
    SHOP_FILLS = [
        (PatternFill("solid", fgColor="1A5276"), PatternFill("solid", fgColor="D4E6F1")),
        (PatternFill("solid", fgColor="145A32"), PatternFill("solid", fgColor="D5F5E3")),
        (PatternFill("solid", fgColor="6E2F1A"), PatternFill("solid", fgColor="FDEBD0")),
        (PatternFill("solid", fgColor="4A235A"), PatternFill("solid", fgColor="F3E6FA")),
    ]

    N_COLS = 10
    W = get_column_letter

    ws.merge_cells(f"A1:{W(N_COLS)}1")
    c = ws["A1"]
    c.value     = f"🚌  送迎ルート表　【{sheet_title}】"
    c.font      = Font(bold=True, size=15, color="FFFFFF", name="メイリオ")
    c.fill      = TITLE_FILL
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    hdr_row = 2
    if date_str:
        ws.merge_cells(f"A2:{W(N_COLS)}2")
        c = ws["A2"]
        c.value     = f"実施日：{date_str}"
        c.font      = Font(bold=True, size=12, color="FFFFFF", name="メイリオ")
        c.fill      = DATE_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _cell_bdr()
        ws.row_dimensions[2].height = 24
        hdr_row = 3

    headers = ["店舗","車両名","運転担当","優先度","順番","利用者氏名","サービス","住所","到着予定","備考"]
    widths  = [10,    20,      14,        6,       6,     16,          14,       34,    10,       14]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=hdr_row, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", size=10, name="メイリオ")
        c.fill      = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _cell_bdr()
        ws.column_dimensions[W(col)].width = w
    ws.row_dimensions[hdr_row].height = 24

    shops = sorted(set(r.shop for r in routes))
    row   = hdr_row + 1

    for si, shop in enumerate(shops):
        dark_fill, light_fill = SHOP_FILLS[si % len(SHOP_FILLS)]
        ws.merge_cells(f"A{row}:{W(N_COLS)}{row}")
        c = ws.cell(row=row, column=1, value=f"　🏠 {shop}")
        c.font      = Font(bold=True, size=12, color="FFFFFF", name="メイリオ")
        c.fill      = dark_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _cell_bdr()
        ws.row_dimensions[row].height = 28
        row += 1

        for route in [r for r in routes if r.shop == shop]:
            dn = route.driver.name     if route.driver else "未定"
            dp = route.driver.priority if route.driver else "-"
            for i, stop in enumerate(route.stops):
                h_, m_ = divmod(stop["arrival_min"], 60)
                fill = WC_FILL if stop["user"].wheelchair else (
                    light_fill if i == 0 else ALT_FILL
                )
                data = [
                    route.shop         if i == 0 else "",
                    route.vehicle.name if i == 0 else "",
                    dn                 if i == 0 else "",
                    dp                 if i == 0 else "",
                    i + 1,
                    stop["user"].name,
                    stop["user"].service_type.value,
                    stop["address"],
                    f"{h_:02d}:{m_:02d}",
                    "♿ 車椅子" if stop["user"].wheelchair else "",
                ]
                for col, val in enumerate(data, 1):
                    c = ws.cell(row=row, column=col, value=val)
                    c.font      = Font(bold=(i == 0), size=10, name="メイリオ")
                    c.alignment = Alignment(
                        horizontal="center" if col in [1,2,3,4,5,9,10] else "left",
                        vertical="center"
                    )
                    c.border = _cell_bdr()
                    if fill:
                        c.fill = fill
                ws.row_dimensions[row].height = 20
                row += 1
        row += 1

    from openpyxl.worksheet.page import PageMargins
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
    ws.print_title_rows = f"1:{hdr_row}"


def _routes_to_rows(routes, trip_label):
    rows = []
    for r in routes:
        for i, stop in enumerate(r.stops):
            h_, m_ = divmod(stop["arrival_min"], 60)
            rows.append({
                "便": trip_label, "店舗": r.shop, "車両名": r.vehicle.name,
                "運転担当": r.driver.name if r.driver else "未定",
                "優先度": r.driver.priority if r.driver else "-",
                "順番": i+1, "氏名": stop["user"].name,
                "サービス": stop["user"].service_type.value,
                "住所": stop["address"], "到着予定": f"{h_:02d}:{m_:02d}",
                "車椅子": "♿" if stop["user"].wheelchair else "",
            })
    return rows


def build_excel_output(pickup_routes, dropoff_routes,
                       target_date: Optional[datetime.date] = None) -> bytes:
    """迎え便・送り便を2シートで出力。"""
    if not OPENPYXL_AVAILABLE:
        rows = _routes_to_rows(pickup_routes, "迎え") + _routes_to_rows(dropoff_routes, "送り")
        buf = io.StringIO()
        pd.DataFrame(rows).to_csv(buf, index=False, encoding="utf-8-sig")
        return buf.getvalue().encode("utf-8-sig")

    wb       = Workbook()
    wb.remove(wb.active)
    date_str = target_date.strftime("%Y/%m/%d") if target_date else ""

    for routes, sname in [(pickup_routes, "迎え便"), (dropoff_routes, "送り便")]:
        ws = wb.create_sheet(title=sname)
        _write_route_sheet(ws, routes, sname, date_str)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ==============================================================
# デモデータ
# ==============================================================

def get_demo_data():
    """デモ用マスタデータ（v7: 4項目の時間枠）。"""
    users = [
        # A店
        User("u1",  "山田 太郎",   "富山市上袋100",   36.720, 137.210, ServiceType.HOUKAGO_DEI, "A店",
             pickup_earliest=870, pickup_latest=540, dropoff_earliest=0, dropoff_latest=1110),
        User("u2",  "鈴木 花子",   "富山市堀川200",   36.695, 137.220, ServiceType.HOUKAGO_DEI, "A店",
             incompatible=["u3"],
             pickup_earliest=870, pickup_latest=570, dropoff_earliest=0, dropoff_latest=1080),
        User("u3",  "田中 一郎",   "富山市婦中300",   36.660, 137.160, ServiceType.A_TYPE,      "A店",
             wheelchair=True, incompatible=["u2"],
             pickup_earliest=0,   pickup_latest=540, dropoff_earliest=0, dropoff_latest=1050),
        User("u4",  "佐藤 愛",     "富山市大沢野400", 36.630, 137.230, ServiceType.B_TYPE,      "A店",
             pickup_earliest=0,   pickup_latest=540, dropoff_earliest=0, dropoff_latest=1050),
        User("u5",  "高橋 健太",   "富山市八尾500",   36.590, 137.270, ServiceType.HOUKAGO_DEI, "A店",
             pickup_earliest=870, pickup_latest=540, dropoff_earliest=0, dropoff_latest=1050),
        # B店
        User("u6",  "渡辺 さくら", "富山市上袋600",   36.725, 137.215, ServiceType.B_TYPE,      "B店",
             pickup_earliest=0,   pickup_latest=540, dropoff_earliest=0, dropoff_latest=1050),
        User("u7",  "伊藤 翔",     "富山市堀川700",   36.700, 137.225, ServiceType.A_TYPE,      "B店",
             pickup_earliest=0,   pickup_latest=540, dropoff_earliest=0, dropoff_latest=1050),
        User("u8",  "中村 みな",   "富山市婦中800",   36.655, 137.155, ServiceType.B_TYPE,      "B店",
             pickup_earliest=0,   pickup_latest=600, dropoff_earliest=0, dropoff_latest=1110),
        User("u9",  "小林 大輝",   "富山市大沢野900", 36.625, 137.235, ServiceType.HOUKAGO_DEI, "B店",
             pickup_earliest=870, pickup_latest=540, dropoff_earliest=0, dropoff_latest=1050),
        User("u10", "加藤 りん",   "富山市八尾1000",  36.585, 137.265, ServiceType.A_TYPE,      "B店",
             pickup_earliest=0,   pickup_latest=540, dropoff_earliest=0, dropoff_latest=1050),
        # C店
        User("u11", "中島 陽斗",   "富山市上袋1100",  36.715, 137.205, ServiceType.HOUKAGO_DEI, "C店",
             pickup_earliest=870, pickup_latest=540, dropoff_earliest=0, dropoff_latest=1050),
        User("u12", "斉藤 みゆ",   "富山市堀川1200",  36.690, 137.215, ServiceType.B_TYPE,      "C店",
             wheelchair=True,
             pickup_earliest=0,   pickup_latest=570, dropoff_earliest=0, dropoff_latest=1080),
    ]
    for u in users:
        u.service_time = 600 if u.wheelchair else 300

    vehicles = [
        Vehicle("v1", "A-1号車（大型）", "large",  7, "A店", True,  36.695, 137.211),
        Vehicle("v2", "A-2号車（普通）", "normal", 4, "A店", False, 36.695, 137.211),
        Vehicle("v3", "B-1号車（大型）", "large",  7, "B店", False, 36.710, 137.200),
        Vehicle("v4", "B-2号車（普通）", "normal", 4, "B店", False, 36.710, 137.200),
        Vehicle("v5", "C-1号車（大型）", "large",  7, "C店", True,  36.680, 137.220),
    ]
    staff = [
        Staff("s1", "林 誠一",   "A店", True,  1, 480,  1140),
        Staff("s2", "森 美咲",   "A店", True,  2, 480,  1020),
        Staff("s3", "池田 裕二", "A店", False, 9, None, None),
        Staff("s4", "宇野 幸子", "B店", True,  1, 480,  1140),
        Staff("s5", "川口 拓也", "B店", True,  2, 540,  1080),
        Staff("s6", "高木 雄介", "C店", True,  1, 480,  1140),
        Staff("s7", "中島 奈々", "C店", True,  3, 600,  1200),
    ]
    return users, vehicles, staff


def build_demo_calendar(users, staff, year: int, month: int) -> dict:
    """
    デモ用月間カレンダーデータ（Excelの数式と同等の動作をPythonで再現）。
    利用者: 月〜金を基本利用日として (pickup_earliest, pickup_latest) を設定。
    スタッフ: 月〜土を基本出勤日として (shift_start, shift_end) を設定。
    """
    _, days_in_month = cal_mod.monthrange(year, month)
    dates = [datetime.date(year, month, d) for d in range(1, days_in_month + 1)]

    # デモ: 月〜金が基本利用日（u1, u2, u5, u9, u11 は学校あり → 早迎え禁止あり）
    SCHOOL_USERS = {"u1", "u2", "u5", "u9", "u11"}

    staff_cal = {}
    for s in staff:
        schedule = {}
        for dt in dates:
            if dt.weekday() == 6:        # 日曜: 休み
                schedule[dt.strftime("%Y-%m-%d")] = None
            elif s.shift_start and s.shift_end:
                schedule[dt.strftime("%Y-%m-%d")] = (s.shift_start, s.shift_end)
            else:
                schedule[dt.strftime("%Y-%m-%d")] = None
        staff_cal[s.name] = schedule

    users_cal = {}
    for u in users:
        schedule = {}
        for dt in dates:
            dow = dt.weekday()
            if dow >= 5:  # 土日: 欠席
                schedule[dt.strftime("%Y-%m-%d")] = None
            else:
                # pickup_earliest: 学校ユーザーは14:30、それ以外は0
                earl = u.pickup_earliest
                last = u.pickup_latest
                schedule[dt.strftime("%Y-%m-%d")] = (earl, last)
        users_cal[u.name] = schedule

    return {"staff": staff_cal, "users": users_cal, "year": year, "month": month}


# ==============================================================
# サンプルExcel生成（v7: 曜日列 + Excel数式）
# ==============================================================

def get_sample_excel(year: Optional[int] = None, month: Optional[int] = None) -> bytes:
    """
    v7対応サンプルExcelを生成。

    シート:
      利用者          - マスタ（A30/B30/C30 + 曜日7列）
      車両            - マスタ
      スタッフ        - マスタ（A10/B6/C6）
      カレンダー_スタッフ - 数式で氏名・シフト時間を自動転記
      カレンダー_利用者   - 数式で曜日×〇 から時間を自動生成
      記入例          - 入力ガイド
    """
    users, vehicles, staff = get_demo_data()

    if year is None or month is None:
        today = datetime.date.today()
        year  = today.year
        month = today.month

    if not OPENPYXL_AVAILABLE:
        buf = io.StringIO()
        pd.DataFrame([{"氏名": u.name, "店舗": u.shop} for u in users]).to_csv(buf, index=False)
        return buf.getvalue().encode("utf-8-sig")

    wb = Workbook()
    wb.remove(wb.active)

    # ── 利用者マスタ（v7: 4時間項目 + 曜日7列）──
    user_headers = [
        "ID", "氏名", "住所", "緯度", "経度",
        "サービス種別", "店舗", "車椅子", "同乗不可ID",
        "早迎え禁止時刻", "迎えリミット", "送り出発可能時刻", "送り到着リミット",
        "月", "火", "水", "木", "金", "土", "日",
    ]
    user_by_shop = {shop: [] for shop in SHOP_LIST}
    for u in users:
        if u.shop in user_by_shop:
            # デモ: 月〜金に〇、土日は空（施設外就労/放デイによって異なる）
            is_school = u.user_id in {"u1", "u2", "u5", "u9", "u11"}
            user_by_shop[u.shop].append({
                "ID":          u.user_id,
                "氏名":        u.name,
                "住所":        u.address,
                "緯度":        u.lat,
                "経度":        u.lng,
                "サービス種別": u.service_type.value,
                "店舗":        u.shop,
                "車椅子":      u.wheelchair,
                "同乗不可ID":  ",".join(u.incompatible),
                "早迎え禁止時刻": min_to_hhmm(u.pickup_earliest) if u.pickup_earliest else "",
                "迎えリミット":   min_to_hhmm(u.pickup_latest),
                "送り出発可能時刻": min_to_hhmm(u.dropoff_earliest) if u.dropoff_earliest else "",
                "送り到着リミット": min_to_hhmm(u.dropoff_latest)   if u.dropoff_latest   else "",
                "月": "〇", "火": "〇", "水": "〇", "木": "〇", "金": "〇",
                "土": "〇" if u.service_type != ServiceType.HOUKAGO_DEI else "",
                "日": "",
            })

    ws_u = wb.create_sheet("利用者")
    _write_master_sheet_v5(
        ws_u, user_by_shop, USER_FRAMES, user_headers,
        title="利用者マスタ　（各店30名枠・空欄行に追記可）", header_color="2C4A6E",
    )

    # ── 車両マスタ ──
    ws_v = wb.create_sheet("車両")
    _write_master_sheet(ws_v, [
        {"ID": v.vehicle_id, "車両名": v.name, "種別コード": v.vehicle_type,
         "定員": v.capacity, "店舗": v.shop, "車椅子対応": v.wheelchair_ok,
         "デポ緯度": v.depot_lat, "デポ経度": v.depot_lng,
         } for v in vehicles
    ], title="車両マスタ", header_color="1A5276")

    # ── スタッフマスタ（A10/B6/C6）──
    staff_headers = ["ID", "氏名", "店舗", "運転可否", "優先度", "出勤時間", "退勤時間"]
    staff_by_shop = {shop: [] for shop in SHOP_LIST}
    for s in staff:
        if s.shop in staff_by_shop:
            staff_by_shop[s.shop].append({
                "ID": s.staff_id, "氏名": s.name, "店舗": s.shop,
                "運転可否": s.can_drive, "優先度": s.priority,
                "出勤時間": min_to_hhmm(s.shift_start) if s.shift_start else "",
                "退勤時間": min_to_hhmm(s.shift_end)   if s.shift_end   else "",
            })

    ws_s = wb.create_sheet("スタッフ")
    _write_master_sheet_v5(
        ws_s, staff_by_shop, STAFF_FRAMES, staff_headers,
        title="スタッフマスタ　（A店10枠 / B店6枠 / C店6枠）", header_color="145A32",
    )

    # ── カレンダーシート（数式埋め込み）──
    _write_calendar_sheet_staff(wb, staff, year, month)
    _write_calendar_sheet_users(wb, users, year, month)

    # ── 記入例 ──
    ws_ex = wb.create_sheet("記入例")
    _write_master_sheet(ws_ex, [
        {"項目": "早迎え禁止時刻",    "例": "14:30",       "説明": "学校が終わる前に迎えに行っても空振りになる時刻（HH:MM）。空欄=制約なし"},
        {"項目": "迎えリミット",      "例": "15:30",       "説明": "これより遅く迎えに行くと問題になる最遅時刻（HH:MM）。必須"},
        {"項目": "送り出発可能時刻",  "例": "17:00",       "説明": "施設を出発できる最早時刻（HH:MM）。空欄=制約なし"},
        {"項目": "送り到着リミット",  "例": "18:30",       "説明": "自宅に届ける最遅時刻（HH:MM）。空欄=制約なし"},
        {"項目": "曜日列（月〜日）",  "例": "〇",           "説明": "その曜日に利用する場合は「〇」を入力。カレンダーに自動反映される"},
        {"項目": "カレンダー上書き",  "例": "10:00-15:00", "説明": "例外の日は直接「HH:MM-HH:MM」を入力して数式を上書きする"},
        {"項目": "出勤時間/退勤時間", "例": "08:00/19:00", "説明": "スタッフのシフト時間（HH:MM）。空欄=制約なし"},
        {"項目": "行の追加方法",      "例": "氏名を入力",   "説明": "空枠の氏名列を記入するとカレンダーに自動反映。枠外は行を挿入して追加可"},
    ], title="記入例・入力ガイド（v7）", header_color="7D3C98")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ==============================================================
# extract_for_date（カレンダーから対象日のデータを抽出）
# ==============================================================

def extract_for_date(calendar_data, all_users, all_staff, target_date) -> tuple:
    """
    カレンダーから指定日の出席者・出勤者を抽出。
    カレンダーに時間情報がある場合はUser/Staffの時間枠を上書きする。
    """
    if calendar_data is None:
        return all_users, all_staff

    import copy
    date_str   = target_date.strftime("%Y-%m-%d")
    staff_cal  = calendar_data.get("staff",  {})
    users_cal  = calendar_data.get("users",  {})

    # スタッフ
    filtered_staff = []
    for s in all_staff:
        schedule = staff_cal.get(s.name, {})
        tw = schedule.get(date_str, "MISSING")
        if tw == "MISSING":
            filtered_staff.append(s)
        elif tw is None:
            pass  # 休み
        else:
            s_new = copy.copy(s)
            s_new.shift_start = tw[0]
            s_new.shift_end   = tw[1]
            filtered_staff.append(s_new)

    # 利用者
    filtered_users = []
    for u in all_users:
        schedule = users_cal.get(u.name, {})
        tw = schedule.get(date_str, "MISSING")
        if tw == "MISSING":
            filtered_users.append(u)
        elif tw is None:
            pass  # 欠席
        else:
            u_new = copy.copy(u)
            # tw[0]=早迎え禁止(0なら制約なし), tw[1]=迎えリミット
            u_new.pickup_earliest = tw[0]   # 0 means no constraint
            u_new.pickup_latest   = tw[1]
            filtered_users.append(u_new)

    return filtered_users, filtered_staff


# ==============================================================
# エディタ用 DataFrame 構築
# ==============================================================

def _build_user_editor_df(all_users, calendar_data, target_date) -> pd.DataFrame:
    """
    利用者エディタ用 DataFrame。
    カレンダーの予定に合わせてチェック状態を自動設定。
    """
    date_str  = target_date.strftime("%Y-%m-%d")
    users_cal = calendar_data.get("users", {}) if calendar_data else {}

    rows = []
    for u in sorted(all_users, key=lambda x: (x.shop, x.name)):
        schedule = users_cal.get(u.name)

        if schedule is None:
            # カレンダー未登録 → デフォルトON
            checked      = True
            earl_disp    = min_to_hhmm(u.pickup_earliest) if u.pickup_earliest else ""
            limit_disp   = min_to_hhmm(u.pickup_latest)
            do_earl_disp = min_to_hhmm(u.dropoff_earliest) if u.dropoff_earliest else ""
            do_last_disp = min_to_hhmm(u.dropoff_latest)   if u.dropoff_latest   else ""
        else:
            tw = schedule.get(date_str, "MISSING")
            if tw == "MISSING" or tw is None:
                checked      = tw == "MISSING"   # MISSING→ON, None→OFF
                earl_disp    = min_to_hhmm(u.pickup_earliest) if u.pickup_earliest else ""
                limit_disp   = min_to_hhmm(u.pickup_latest)
                do_earl_disp = min_to_hhmm(u.dropoff_earliest) if u.dropoff_earliest else ""
                do_last_disp = min_to_hhmm(u.dropoff_latest)   if u.dropoff_latest   else ""
            else:
                checked      = True
                earl_disp    = min_to_hhmm(tw[0]) if tw[0] > 0 else ""
                limit_disp   = min_to_hhmm(tw[1])
                do_earl_disp = min_to_hhmm(u.dropoff_earliest) if u.dropoff_earliest else ""
                do_last_disp = min_to_hhmm(u.dropoff_latest)   if u.dropoff_latest   else ""

        rows.append({
            "出席":          checked,
            "店舗":          u.shop,
            "氏名":          u.name,
            "サービス":      u.service_type.value,
            "車椅子":        "♿" if u.wheelchair else "",
            "早迎え禁止":    earl_disp,
            "迎えリミット":  limit_disp,
            "送り出発可能":  do_earl_disp,
            "送り到着リミット": do_last_disp,
            "_uid":          u.user_id,
        })
    return pd.DataFrame(rows)


def _build_staff_editor_df(all_staff, calendar_data, target_date) -> pd.DataFrame:
    """スタッフエディタ用 DataFrame。カレンダー連動チェック。"""
    date_str  = target_date.strftime("%Y-%m-%d")
    staff_cal = calendar_data.get("staff", {}) if calendar_data else {}

    rows = []
    for s in sorted(all_staff, key=lambda x: (x.shop, x.priority)):
        schedule = staff_cal.get(s.name)

        if schedule is None:
            checked  = True
            ss_disp  = min_to_hhmm(s.shift_start) if s.shift_start else "08:00"
            se_disp  = min_to_hhmm(s.shift_end)   if s.shift_end   else "19:00"
        else:
            tw = schedule.get(date_str, "MISSING")
            if tw == "MISSING":
                checked = True
                ss_disp = min_to_hhmm(s.shift_start) if s.shift_start else "08:00"
                se_disp = min_to_hhmm(s.shift_end)   if s.shift_end   else "19:00"
            elif tw is None:
                checked = False
                ss_disp = min_to_hhmm(s.shift_start) if s.shift_start else "08:00"
                se_disp = min_to_hhmm(s.shift_end)   if s.shift_end   else "19:00"
            else:
                checked = True
                ss_disp = min_to_hhmm(tw[0])
                se_disp = min_to_hhmm(tw[1])

        rows.append({
            "出勤":     checked,
            "店舗":     s.shop,
            "氏名":     s.name,
            "優先度":   s.priority,
            "運転可否": s.can_drive,
            "出勤時間": ss_disp,
            "退勤時間": se_disp,
            "_sid":     s.staff_id,
        })
    return pd.DataFrame(rows)


def _reconstruct_users_from_editor(edited_df, all_users) -> list:
    """data_editor の編集結果から User リストを再構築する。"""
    import copy
    user_map = {u.user_id: u for u in all_users}
    result   = []

    for _, row in edited_df.iterrows():
        if not row.get("出席", False):
            continue
        uid  = str(row.get("_uid", ""))
        orig = user_map.get(uid)
        if orig is None:
            continue

        u_new = copy.copy(orig)

        def _rt(key, default):
            v = str(row.get(key, "")).strip()
            if not v or v == "nan":
                return default
            return hhmm_to_min(v, default)

        u_new.pickup_earliest  = _rt("早迎え禁止",     orig.pickup_earliest)
        u_new.pickup_latest    = _rt("迎えリミット",   orig.pickup_latest)
        u_new.dropoff_earliest = _rt("送り出発可能",   orig.dropoff_earliest)
        u_new.dropoff_latest   = _rt("送り到着リミット", orig.dropoff_latest)
        result.append(u_new)

    return result


def _reconstruct_staff_from_editor(edited_df, all_staff) -> list:
    """data_editor の編集結果から Staff リストを再構築する。"""
    import copy
    staff_map = {s.staff_id: s for s in all_staff}
    result    = []

    for _, row in edited_df.iterrows():
        if not row.get("出勤", False):
            continue
        sid  = str(row.get("_sid", ""))
        orig = staff_map.get(sid)
        if orig is None:
            continue

        s_new = copy.copy(orig)
        s_new.can_drive = bool(row.get("運転可否", orig.can_drive))

        ss_raw = str(row.get("出勤時間", "")).strip()
        se_raw = str(row.get("退勤時間", "")).strip()
        if ss_raw and ss_raw != "nan":
            ss = hhmm_to_min(ss_raw, -1)
            s_new.shift_start = ss if ss >= 0 else orig.shift_start
        if se_raw and se_raw != "nan":
            se = hhmm_to_min(se_raw, -1)
            s_new.shift_end = se if se >= 0 else orig.shift_end

        result.append(s_new)

    return result


# ==============================================================
# 結果テーブル → DataFrame
# ==============================================================

def routes_to_dataframe(routes) -> pd.DataFrame:
    rows = []
    for route in routes:
        dn = route.driver.name     if route.driver else "未定"
        dp = route.driver.priority if route.driver else "-"
        for i, stop in enumerate(route.stops):
            h_, m_ = divmod(stop["arrival_min"], 60)
            rows.append({
                "店舗":     route.shop,
                "車両名":   route.vehicle.name,
                "運転担当": dn,
                "優先度":   dp,
                "順番":     i + 1,
                "氏名":     stop["user"].name,
                "サービス": stop["user"].service_type.value,
                "住所":     stop["address"],
                "到着予定": f"{h_:02d}:{m_:02d}",
                "車椅子":   "♿" if stop["user"].wheelchair else "",
            })
    return pd.DataFrame(rows) if rows else pd.DataFrame()


# ==============================================================
# 地図描画
# ==============================================================

SHOP_MAP_COLORS = ["blue", "green", "red", "purple", "orange", "darkblue"]


def render_map(routes):
    if not FOLIUM_AVAILABLE:
        st.info("📦 `folium` と `streamlit-folium` をインストールすると地図が表示されます")
        return
    if not routes:
        return

    all_lats = [s["lat"] for r in routes for s in r.stops]
    all_lngs = [s["lng"] for r in routes for s in r.stops]
    center   = [sum(all_lats)/len(all_lats), sum(all_lngs)/len(all_lngs)]
    m        = folium.Map(location=center, zoom_start=12, tiles="CartoDB positron")

    shops = sorted(set(r.shop for r in routes))
    sc    = {s: SHOP_MAP_COLORS[i % len(SHOP_MAP_COLORS)] for i, s in enumerate(shops)}

    for route in routes:
        color  = sc[route.shop]
        depot  = [route.vehicle.depot_lat, route.vehicle.depot_lng]
        coords = [depot] + [[s["lat"], s["lng"]] for s in route.stops] + [depot]

        folium.Marker(depot, tooltip=f"🏠 {route.shop} デポ",
                      icon=folium.Icon(color=color, icon="home", prefix="fa")).add_to(m)
        folium.PolyLine(coords, color=color, weight=3, opacity=0.75,
                        tooltip=f"{route.shop} - {route.vehicle.name}").add_to(m)

        for i, stop in enumerate(route.stops):
            h_, mn_ = divmod(stop["arrival_min"], 60)
            wc      = "♿ " if stop["user"].wheelchair else ""
            popup   = (
                f"<b>{i+1}. {stop['user'].name}</b><br>"
                f"{wc}{stop['address']}<br>"
                f"🕐 {h_:02d}:{mn_:02d}<br>"
                f"🚗 {route.vehicle.name} ({route.shop})"
            )
            folium.CircleMarker(
                location=[stop["lat"], stop["lng"]],
                radius=8, color=color, fill=True, fill_opacity=0.85,
                tooltip=f"{i+1}. {stop['user'].name}",
                popup=folium.Popup(popup, max_width=240),
            ).add_to(m)

    st_folium(m, width=None, height=480, returned_objects=[])


# ==============================================================
# Streamlit UI メイン
# ==============================================================

def main():

    # ヘッダー
    st.markdown("""
    <div class="hero fade-up">
      <div class="v-badge">VERSION 7</div>
      <h1>送迎ルート最適化システム</h1>
      <p>
        放課後等デイサービス・就労継続支援A型/B型 対応　｜　3店舗混載禁止　｜　月間カレンダー連動<br>
        曜日別基本設定・カレンダー数式自動転記・UI上での直前微調整
      </p>
    </div>
    """, unsafe_allow_html=True)

    # ================================================================
    # サイドバー
    # ================================================================
    with st.sidebar:
        algo_color = "#27AE60" if ORTOOLS_AVAILABLE else "#E67E22"
        algo_label = "OR-Tools（高精度VRP）" if ORTOOLS_AVAILABLE else "グリーディ（高速）"
        st.markdown(f"""
        <div class="sidebar-section">
          <h4>⚙️ システム状態</h4>
          <div style="font-size:12px;color:{algo_color};font-weight:700;">● {algo_label}</div>
          <div style="font-size:11px;color:#888;margin-top:4px;">
            {"OR-Tools インストール済み" if ORTOOLS_AVAILABLE else "pip install ortools で高精度モードに"}
          </div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown('<div class="sidebar-section"><h4>📅 送迎実施日</h4>', unsafe_allow_html=True)
        target_date = st.date_input("送迎実施日", value=datetime.date.today(),
                                    label_visibility="collapsed")
        st.markdown(f'<div class="cal-date-badge">{target_date.strftime("%Y年 %m月 %d日 (%a)")}</div>',
                    unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

        st.markdown('<div class="sidebar-section"><h4>⏰ デフォルト時刻</h4>', unsafe_allow_html=True)
        st.caption("Excelが空欄の場合のフォールバック値")
        st.markdown("**迎え便**")
        c1, c2 = st.columns(2)
        with c1:
            pu_sh = st.number_input("出発 時", 5,  12, 8,  key="pu_sh")
            pu_sm = st.number_input("出発 分", 0,  55, 0,  step=5, key="pu_sm")
        with c2:
            pu_lh = st.number_input("リミット 時", 6, 13, 9,  key="pu_lh")
            pu_lm = st.number_input("リミット 分", 0, 55, 0,  step=5, key="pu_lm")
        st.markdown("**送り便**")
        c3, c4 = st.columns(2)
        with c3:
            do_sh = st.number_input("出発 時", 15, 20, 17, key="do_sh")
            do_sm = st.number_input("出発 分", 0,  55, 0,  step=5, key="do_sm")
        with c4:
            do_lh = st.number_input("リミット 時", 16, 22, 19, key="do_lh")
            do_lm = st.number_input("リミット 分", 0,  55, 0,  step=5, key="do_lm")

        pu_start = pu_sh * 60 + pu_sm
        pu_limit = pu_lh * 60 + pu_lm
        do_start = do_sh * 60 + do_sm
        do_limit = do_lh * 60 + do_lm
        st.markdown("</div>", unsafe_allow_html=True)

        st.markdown('<div class="sidebar-section"><h4>📥 テンプレート</h4>', unsafe_allow_html=True)
        if OPENPYXL_AVAILABLE:
            st.download_button(
                "サンプルExcel（v7対応）をダウンロード",
                data=get_sample_excel(target_date.year, target_date.month),
                file_name=f"送迎マスタ_{target_date.year}{target_date.month:02d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.caption("シート: 利用者/車両/スタッフ/カレンダー_スタッフ/カレンダー_利用者")
        st.markdown("</div>", unsafe_allow_html=True)

    # ================================================================
    # STEP 1: データ読み込み
    # ================================================================
    step_header(1, "Excelを読み込む",
                "1ヶ月分の利用予定・シフトが入ったExcelをアップロード")

    col_up, col_demo = st.columns([3, 1])
    with col_up:
        uploaded = st.file_uploader("Excelファイル（.xlsx）", type=["xlsx", "xls"],
                                    help="サイドバーからサンプルExcelをDL")
        if uploaded:
            try:
                with st.spinner("読み込み中..."):
                    users, vehicles, staff, cal_data = parse_excel_upload(
                        uploaded, pu_limit, do_limit
                    )
                st.session_state.update({"users": users, "vehicles": vehicles,
                                         "staff": staff, "calendar": cal_data})
                cal_msg = ""
                if cal_data:
                    sc = cal_data.get("staff", {}); uc = cal_data.get("users", {})
                    yr = cal_data.get("year", ""); mo = cal_data.get("month", "")
                    cal_msg = f"　月間カレンダー **{yr}年{mo}月** (スタッフ{len(sc)}名/利用者{len(uc)}名)"
                st.success(
                    f"✅ 読み込み完了　利用者 **{len(users)}名** / 車両 **{len(vehicles)}台** / スタッフ **{len(staff)}名**{cal_msg}"
                )
            except Exception as e:
                st.error(f"❌ 読み込み失敗: {e}")
                st.info("シート名・列名がサンプルExcelと一致しているか確認してください")

    with col_demo:
        st.markdown("<br><br>", unsafe_allow_html=True)
        if st.button("🎯 デモデータで試す", use_container_width=True, type="secondary"):
            users, vehicles, staff = get_demo_data()
            cal_data = build_demo_calendar(users, staff, target_date.year, target_date.month)
            st.session_state.update({"users": users, "vehicles": vehicles,
                                     "staff": staff, "calendar": cal_data})
            st.success("✅ デモデータ読み込み完了")

    if "users" not in st.session_state:
        st.info("👆 Excelをアップロードするか「デモデータで試す」をクリックしてください")
        return

    all_users     = st.session_state["users"]
    all_vehicles  = st.session_state["vehicles"]
    all_staff     = st.session_state["staff"]
    calendar_data = st.session_state.get("calendar")

    with st.expander("📋 マスタデータを確認する", expanded=False):
        t1, t2, t3, t4 = st.tabs(["👶 利用者", "🚗 車両", "👤 スタッフ", "📅 カレンダー"])
        with t1:
            st.dataframe(pd.DataFrame([{
                "氏名": u.name, "店舗": u.shop, "サービス": u.service_type.value,
                "住所": u.address, "車椅子": "♿" if u.wheelchair else "",
                "早迎え禁止": min_to_hhmm(u.pickup_earliest) if u.pickup_earliest else "なし",
                "迎えリミット": min_to_hhmm(u.pickup_latest),
                "送り出発可能": min_to_hhmm(u.dropoff_earliest) if u.dropoff_earliest else "なし",
                "送り到着リミット": min_to_hhmm(u.dropoff_latest) if u.dropoff_latest else "なし",
            } for u in all_users]), use_container_width=True, hide_index=True)
        with t2:
            st.dataframe(pd.DataFrame([{
                "車両名": v.name, "店舗": v.shop, "種別": v.vehicle_type,
                "定員": v.capacity, "車椅子対応": "✅" if v.wheelchair_ok else "✗",
            } for v in all_vehicles]), use_container_width=True, hide_index=True)
        with t3:
            st.dataframe(pd.DataFrame([{
                "氏名": s.name, "店舗": s.shop,
                "運転": "✅" if s.can_drive else "❌", "優先度": s.priority,
                "出勤": min_to_hhmm(s.shift_start) if s.shift_start else "終日",
                "退勤": min_to_hhmm(s.shift_end)   if s.shift_end   else "終日",
            } for s in all_staff]), use_container_width=True, hide_index=True)
        with t4:
            if calendar_data:
                sc = calendar_data.get("staff", {}); uc = calendar_data.get("users", {})
                yr = calendar_data.get("year", "?"); mo = calendar_data.get("month", "?")
                st.caption(f"📅 {yr}年{mo}月　スタッフ **{len(sc)}名**分　利用者 **{len(uc)}名**分")
                col_s, col_u = st.columns(2)
                with col_s:
                    st.markdown("**スタッフ（最初の7日）**")
                    if sc:
                        nm = next(iter(sc))
                        rows = [{"日付": d, "時間": f"{min_to_hhmm(v[0])}〜{min_to_hhmm(v[1])}" if v else "休み"}
                                for d, v in list(sc[nm].items())[:7]]
                        st.caption(f"{nm}")
                        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
                with col_u:
                    st.markdown("**利用者（最初の7日）**")
                    if uc:
                        nm = next(iter(uc))
                        rows = [{"日付": d, "時間": f"{min_to_hhmm(v[0])}〜{min_to_hhmm(v[1])}" if v else "欠席"}
                                for d, v in list(uc[nm].items())[:7]]
                        st.caption(f"{nm}")
                        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
            else:
                st.info("月間カレンダーシートが読み込まれていません")

    st.divider()

    # ================================================================
    # STEP 2: 日付選択・カレンダー自動チェック・微調整
    # ================================================================
    step_header(
        2,
        f"出欠・シフトを確認して微調整する　{target_date.strftime('(%m/%d)')}",
        "カレンダーから自動チェック → そのまま実行 or 変更して微調整",
    )

    if calendar_data:
        sc = calendar_data.get("staff", {}); uc = calendar_data.get("users", {})
        date_str = target_date.strftime("%Y-%m-%d")
        auto_u = sum(1 for u in all_users
                     if uc.get(u.name, {}).get(date_str, "MISSING") not in (None,))
        auto_s = sum(1 for s in all_staff
                     if sc.get(s.name, {}).get(date_str, "MISSING") not in (None,))
        st.info(
            f"📅 **{target_date.strftime('%Y年%m月%d日')}** のカレンダーを参照しました。"
            f"　利用者 **{auto_u}名** / スタッフ **{auto_s}名** が自動チェックONです。"
            f"　変更がある場合は下の表を直接編集してください。"
        )
    else:
        st.info("月間カレンダーが未読み込みのため、全員チェックONで表示しています。")

    # ── 利用者エディタ ──
    st.markdown("#### 👶 利用者")
    st.caption(
        "✅ チェック = 本日送迎対象。時間項目はHH:MM形式で直接書き換えられます（空欄=制約なし）。"
    )

    user_df = _build_user_editor_df(all_users, calendar_data, target_date)
    edited_user_df = st.data_editor(
        user_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "出席":           st.column_config.CheckboxColumn("出席",       width="small"),
            "店舗":           st.column_config.TextColumn("店舗",           width="small"),
            "氏名":           st.column_config.TextColumn("氏名",           width="medium"),
            "サービス":       st.column_config.TextColumn("サービス",       width="small"),
            "車椅子":         st.column_config.TextColumn("♿",             width="small"),
            "早迎え禁止":     st.column_config.TextColumn(
                "早迎え禁止", width="small",
                help="これより前に迎えに行っても空振り（HH:MM）。空欄=制約なし",
            ),
            "迎えリミット":   st.column_config.TextColumn(
                "迎えリミット", width="small",
                help="これより遅く迎えに行くと問題になる最遅時刻（HH:MM）",
            ),
            "送り出発可能":   st.column_config.TextColumn(
                "送り出発可能", width="small",
                help="施設を出発できる最早時刻（HH:MM）。空欄=制約なし",
            ),
            "送り到着リミット": st.column_config.TextColumn(
                "送り到着リミット", width="small",
                help="自宅に届ける最遅時刻（HH:MM）。空欄=制約なし",
            ),
            "_uid": st.column_config.TextColumn("_uid", disabled=True),
        },
        column_order=[
            "出席", "店舗", "氏名", "サービス", "車椅子",
            "早迎え禁止", "迎えリミット", "送り出発可能", "送り到着リミット",
        ],
        disabled=["店舗", "氏名", "サービス", "車椅子"],
        key="user_editor_v7",
    )

    attending_users = _reconstruct_users_from_editor(edited_user_df, all_users)

    # ── スタッフエディタ ──
    st.markdown("#### 👤 スタッフ")
    st.caption(
        "✅ チェック = 本日出勤。「運転可否」「出勤/退勤時間」を直接変更できます。"
    )

    staff_df = _build_staff_editor_df(all_staff, calendar_data, target_date)
    edited_staff_df = st.data_editor(
        staff_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "出勤":     st.column_config.CheckboxColumn("出勤",   width="small"),
            "店舗":     st.column_config.TextColumn("店舗",       width="small"),
            "氏名":     st.column_config.TextColumn("氏名",       width="medium"),
            "優先度":   st.column_config.NumberColumn("優先度",   width="small"),
            "運転可否": st.column_config.CheckboxColumn(
                "運転", width="small",
                help="今日だけ運転不可にする場合はチェックを外してください",
            ),
            "出勤時間": st.column_config.TextColumn(
                "出勤時間", width="small", help="HH:MM形式",
            ),
            "退勤時間": st.column_config.TextColumn(
                "退勤時間", width="small", help="HH:MM形式",
            ),
            "_sid": st.column_config.TextColumn("_sid", disabled=True),
        },
        column_order=["出勤", "店舗", "氏名", "優先度", "運転可否", "出勤時間", "退勤時間"],
        disabled=["店舗", "氏名", "優先度"],
        key="staff_editor_v7",
    )

    attending_staff  = _reconstruct_staff_from_editor(edited_staff_df, all_staff)
    active_vehicles  = [v for v in all_vehicles if any(s.shop == v.shop for s in attending_staff)]

    # サマリー
    n_attend = len(attending_users)
    n_absent  = len(all_users) - n_attend
    n_staff   = len(attending_staff)
    n_no_drv  = sum(1 for s in attending_staff if not s.can_drive)

    col_a, col_b, col_c, col_d = st.columns(4)
    col_a.metric("本日 出席利用者",  f"{n_attend} 名")
    col_b.metric("欠席",             f"{n_absent} 名",
                 delta=f"-{n_absent}" if n_absent else None, delta_color="inverse")
    col_c.metric("出勤スタッフ",     f"{n_staff} 名")
    col_d.metric("稼働車両",         f"{len(active_vehicles)} 台")
    if n_no_drv:
        st.caption(f"※ 出勤スタッフのうち {n_no_drv}名 は運転不可として配車から除外されます")

    st.divider()

    # ================================================================
    # STEP 3: 最適化実行 → 結果表示 → ダウンロード
    # ================================================================
    step_header(3, "最適化して送迎ルートを確認する",
                "編集済みの内容でVRPを実行し、Excelでダウンロードできます")

    if n_attend == 0:
        st.warning("出席利用者が0名です。STEP 2 で確認してください。")
        return

    checker = ConstraintChecker()
    errors  = checker.validate(attending_users, active_vehicles, attending_staff)
    for e in errors:
        st.error(e)

    run_btn = st.button(
        f"🚀　{target_date.strftime('%m/%d')} の送迎ルートを最適化する"
        f"　（利用者{n_attend}名 / スタッフ{n_staff}名）",
        disabled=bool(errors),
        type="primary",
        use_container_width=True,
    )

    if run_btn:
        with st.spinner("🔄 迎え便・送り便を同時最適化中..."):
            pickup_routes  = run_all_shops(
                attending_users, active_vehicles, attending_staff,
                TripType.PICKUP, pu_start, pu_limit,
            )
            dropoff_routes = run_all_shops(
                attending_users, active_vehicles, attending_staff,
                TripType.DROPOFF, do_start, do_limit,
            )
        st.session_state.update({
            "pickup_routes":  pickup_routes,
            "dropoff_routes": dropoff_routes,
            "result_date":    target_date,
        })
        st.success(
            f"✅ 最適化完了！　迎え便 **{len(pickup_routes)}** ルート　"
            f"送り便 **{len(dropoff_routes)}** ルート"
        )

    if "pickup_routes" not in st.session_state:
        st.info("👆 上のボタンを押すとルートが表示されます")
        return

    pickup_routes  = st.session_state["pickup_routes"]
    dropoff_routes = st.session_state["dropoff_routes"]
    result_date    = st.session_state.get("result_date")
    all_routes     = pickup_routes + dropoff_routes

    if not all_routes:
        st.error("ルートを生成できませんでした。データを確認してください。")
        return

    # メトリクス
    total_pu   = sum(len(r.stops) for r in pickup_routes)
    total_do   = sum(len(r.stops) for r in dropoff_routes)
    wc_count   = sum(1 for r in pickup_routes for s in r.stops if s["user"].wheelchair)
    shops_used = sorted(set(r.shop for r in all_routes))
    metric_row([
        (total_pu,        "名", "迎え便 乗車合計"),
        (total_do,        "名", "送り便 乗車合計"),
        (wc_count,        "名", "車椅子対応"),
        (len(shops_used), "店", "稼働店舗数"),
    ])

    # 制約検証
    with st.expander("🔍 制約条件 検証サマリー", expanded=False):
        forbidden = checker.get_forbidden_pairs(attending_users)
        all_ok    = True
        for label, routes in [("迎え便", pickup_routes), ("送り便", dropoff_routes)]:
            st.markdown(f"**{label}**")
            for route in routes:
                uin     = [s["user"] for s in route.stops]
                ok_cap  = len(uin) <= route.vehicle.capacity
                ok_wc   = not (any(u.wheelchair for u in uin) and not route.vehicle.wheelchair_ok)
                ok_incp = not any(
                    tuple(sorted([u1.user_id, u2.user_id])) in forbidden
                    for i, u1 in enumerate(uin) for u2 in uin[i+1:]
                )
                ok_drv  = route.driver is not None and route.driver.can_drive
                ok_shop = all(u.shop == route.shop for u in uin)
                all_ok  = all_ok and all([ok_cap, ok_wc, ok_incp, ok_drv, ok_shop])

                def st_icon(ok):
                    return '<span class="ok">✅</span>' if ok else '<span class="fail">❌</span>'

                dn = route.driver.name if route.driver else "未定"
                st.markdown(
                    f"　**{route.shop} - {route.vehicle.name}**"
                    f"（{len(uin)}/{route.vehicle.capacity}名）　"
                    f"定員:{st_icon(ok_cap)} 車椅子:{st_icon(ok_wc)} "
                    f"同乗不可:{st_icon(ok_incp)} 混載:{st_icon(ok_shop)} 運転:{dn} {st_icon(ok_drv)}",
                    unsafe_allow_html=True,
                )
        if all_ok:
            st.success("🎉 全制約条件クリア！")

    # 結果テーブル
    st.markdown("#### 📋 送迎ルート一覧")
    tab_pu, tab_do = st.tabs(["▶ 迎え便", "◀ 送り便"])
    col_cfg = {
        "順番":     st.column_config.NumberColumn(width="small"),
        "到着予定": st.column_config.TextColumn(width="small"),
        "車椅子":   st.column_config.TextColumn(width="small"),
        "優先度":   st.column_config.NumberColumn(width="small"),
    }
    with tab_pu:
        df_pu = routes_to_dataframe(pickup_routes)
        st.dataframe(df_pu, use_container_width=True, hide_index=True, column_config=col_cfg) \
            if not df_pu.empty else st.info("迎え便のルートがありません")
    with tab_do:
        df_do = routes_to_dataframe(dropoff_routes)
        st.dataframe(df_do, use_container_width=True, hide_index=True, column_config=col_cfg) \
            if not df_do.empty else st.info("送り便のルートがありません")

    # タイムライン
    with st.expander("🕐 タイムラインプレビュー（迎え便）", expanded=False):
        for route in sorted(pickup_routes, key=lambda r: r.shop):
            dn = route.driver.name if route.driver else "未定"
            st.markdown(f"**{route.shop} - {route.vehicle.name}**　運転: {dn}")
            html = ""
            for stop in route.stops:
                h_, m_ = divmod(stop["arrival_min"], 60)
                wc     = "♿ " if stop["user"].wheelchair else ""
                lim    = min_to_hhmm(stop["user"].pickup_latest)
                html  += (
                    f'<div class="timeline-item">'
                    f'<div class="timeline-dot"></div>'
                    f'<div class="timeline-time">{h_:02d}:{m_:02d}</div>'
                    f'<div><div class="timeline-name">{wc}{stop["user"].name}</div>'
                    f'<div class="timeline-detail">{stop["address"]}　迎えリミット: {lim}</div>'
                    f'</div></div>'
                )
            st.markdown(html, unsafe_allow_html=True)

    # ダウンロード
    st.markdown("<br>", unsafe_allow_html=True)
    col_dl, _ = st.columns([1, 2])
    with col_dl:
        excel_bytes = build_excel_output(pickup_routes, dropoff_routes, result_date)
        date_fname  = result_date.strftime("%Y%m%d") if result_date else "送迎ルート"
        ext  = "xlsx" if OPENPYXL_AVAILABLE else "csv"
        mime = ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                if OPENPYXL_AVAILABLE else "text/csv")
        st.download_button(
            label=f"📥 送迎ルート表をダウンロード（{date_fname}）",
            data=excel_bytes,
            file_name=f"送迎ルート_{date_fname}.{ext}",
            mime=mime,
            use_container_width=True,
        )
        st.caption("迎え便・送り便を1ファイルの2シートで出力します（A4横向き）")

    st.divider()

    # 地図
    st.markdown("#### 🗺️ 送迎ルートマップ")
    tab_map1, tab_map2 = st.tabs(["▶ 迎え便", "◀ 送り便"])
    with tab_map1:
        render_map(pickup_routes)
    with tab_map2:
        render_map(dropoff_routes)


if __name__ == "__main__":
    main()
